"""
=============================================================================
  CONCALL TRANSCRIPT ANALYZER FOR NSE/BSE LISTED COMPANIES
  ============================================================
  Author: Auto-generated modular pipeline script
  Purpose: Download, parse, and analyze earnings call transcripts
           for Indian-listed companies with expandable filtering.
=============================================================================

PIPELINE OVERVIEW:
  1. Fetch broad company universe  (NSE/BSE)
  2. Apply financial filters       (market cap, PE, sector, etc.)
  3. Download concall transcripts  (BSE XBRL / Screener.in)
  4. Parse PDF text
  5. Score against keyword dict
  6. Export prioritized CSV output

HOW TO ADD A NEW FILTER:
  - Write a function: filter_by_<your_criterion>(df) -> pd.DataFrame
  - Append it to FILTER_PIPELINE list near the bottom of the file.
  - That's it — the pipeline will pick it up automatically.
=============================================================================
"""
from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import numpy as np
import os
import re
import csv
import time
import logging
import random
import warnings
import requests
import io
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION  ← edit these variables freely
# ─────────────────────────────────────────────────────────────────────────────

# ── Market cap filter (in Crores INR) ────────────────────────────────────────
MIN_MARKET_CAP_CRORES = 5_000          # Only companies ≥ this pass through

# ── Keyword priority thresholds ──────────────────────────────────────────────
HIGH_PRIORITY_THRESHOLD   = 5          # ≥5 categories matched  → High
MEDIUM_PRIORITY_THRESHOLD = 4          # 4-5 categories matched → Medium
LOW_PRIORITY_THRESHOLD    = 2          # 2-3 categories matched → Low
                                       # <2 categories          → filtered out

# ── Output settings ──────────────────────────────────────────────────────────
OUTPUT_DIR    = Path("concall_output")
TRANSCRIPTS_DIR = OUTPUT_DIR / "transcripts"
OUTPUT_CSV    = OUTPUT_DIR / f"concall_analysis_{datetime.today().strftime('%Y%m%d')}.csv"

# ── Scraping politeness ──────────────────────────────────────────────────────
SLEEP_MIN_SECS = 3                     # Min sleep between HTTP requests
SLEEP_MAX_SECS = 6                     # Max sleep between HTTP requests

# ─────────────────────────────────────────────────────────────────────────────
# KEYWORD DICTIONARY
# ─────────────────────────────────────────────────────────────────────────────
# Each key is a CATEGORY (counts as 1 point max, regardless of how many
# synonyms are found).  Add/remove categories or synonyms freely.
# ─────────────────────────────────────────────────────────────────────────────

KEYWORDS_DICT = {
    "High Growth / Topline": [
        "topline growth", "revenue growth", "sales growth"
    ],
    "Order Visibility": [
        "order book", "new orders", "order inflow"
    ],
    "Margin Expansion": [
        "margin improvement", "ebit margin growth", "ebit margin expansion", "ebit margins growth", "ebit margins expansion",
        "ebitda margin growth", "ebitda margin expansion", "ebitda margins growth", "ebitda margins expansion",
        "better product mix", "premiumization"
    ],
    "Capacity Expansion": [
        "capacity expansion", "greenfield expansion", "brownfield expansion", "greenfield", 
        "brownfield", "new plant", "new facility", "capex for capacity", "capacity addition"
    ],
    "Integration & Efficiency": [
        "debottlenecking", "backward integration", "forward integration",
        "economies of scale", "improved operational efficiency"
    ],
    "Market Share": [
        "gaining market share", "improving market share", 
        "gaining wallet share", "improving wallet share"
    ],
    "Deleveraging": [
        "debt repayment", "deleveraging", "improved debt equity ratio", "debt replacement",
        "better rate of borrowing", "better rate of refinancing"
    ],
    "Cash Generation": [
        "net debt negative", "net cash positive", "fcf positive", "free cash flow positive"
    ],
    "Working Capital": [
        "working capital improvement", "working capital days reduced", 
        "cash conversion cycle improvement", "cash conversion cycle days reduced"
    ],
    "Return Ratios": [
        "roe improved", "roce improved", "roic improved", "roi improved",
        "return on equity improved", "return on capital employed improved", 
        "return on invested capital improved", "return on investment improved"
    ],
    "Corporate Action": [
        "acquisition", "merger", "demerger", "restructuring"
    ]
}

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(OUTPUT_DIR / "run.log" if OUTPUT_DIR.exists()
                            else Path("concall_run.log")),
    ]
)
log = logging.getLogger("ConcallAnalyzer")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — FETCH COMPANY UNIVERSE
# ─────────────────────────────────────────────────────────────────────────────

def fetch_nse_universe() -> pd.DataFrame:
    """
    Download the master list of ALL listed equities on the NSE.
    """
    log.info("Fetching the complete NSE master company universe …")

    try:
        # The official NSE master list of all listed equities
        csv_url = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
        
        # We must spoof a browser user-agent, otherwise the NSE server blocks the download
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        
        s = requests.Session()
        s.headers.update(headers)
        r = s.get(csv_url, timeout=15)
        r.raise_for_status()
        
        # Read the downloaded CSV content
        df = pd.read_csv(io.StringIO(r.text))

        df.columns = df.columns.str.strip()
        
        # The CSV has columns like 'SYMBOL' and 'NAME OF COMPANY'
        df = df.rename(columns={
            "SYMBOL": "symbol",
            "NAME OF COMPANY": "company_name",
        })
        
        # Filter for standard equities (EQ) and book-entry (BE) to exclude bonds/ETFs
        df = df[df['SERIES'].isin(['EQ', 'BE'])].copy()
        
        # Initialize market cap to None so yfinance can fetch them all in the next step
        df["market_cap_crores"] = None          
        
        log.info(f"  ✓ Full NSE universe fetched: {len(df)} companies")
        return df[["symbol", "company_name", "market_cap_crores"]]

    except Exception as e:
        log.error(f"  Failed to fetch NSE master list: {e}")
        return pd.DataFrame(columns=["symbol", "company_name", "market_cap_crores"])


def enrich_market_cap(df: pd.DataFrame) -> pd.DataFrame:
    """
    For rows where market_cap_crores is null, fetch it via yfinance.
    This runs only for the fallback path; skip if data already present.
    """
    if df["market_cap_crores"].notna().all():
        return df

    log.info("Enriching market cap data via yfinance …")
    try:
        import yfinance as yf
    except ImportError:
        log.warning("  yfinance not installed. Run: pip install yfinance")
        return df

    caps = []
    for sym in df["symbol"]:
        try:
            ticker = yf.Ticker(f"{sym}.NS")
            info   = ticker.fast_info
            mcap   = getattr(info, "market_cap", None)
            caps.append(mcap / 1e7 if mcap else None)   # ₹ → Crores
        except Exception:
            caps.append(None)
        time.sleep(0.2)                                  # light rate-limit

    df["market_cap_crores"] = caps
    log.info(f"  ✓ Market cap enriched for {df['market_cap_crores'].notna().sum()} symbols")
    return df


def fetch_nifty50_universe() -> pd.DataFrame:
    """
    Return the Nifty 50 company list as a ready-to-use DataFrame.

    Used instead of fetch_nse_universe() + enrich_market_cap() when you want
    to test the scoring logic on just 50 companies without waiting for the
    full 2000-company yfinance enrichment loop.

    Market caps are approximate (order-of-magnitude) — sufficient for the
    market cap filter to pass all Nifty 50 companies through (they all exceed
    5,000 Cr comfortably). Screener uses NSE symbols directly, so no BSE
    code resolution is needed.

    To switch back to the full NSE run, replace fetch_nifty50_universe() in
    main() with fetch_nse_universe() + enrich_market_cap().
    """
    log.info("TEST MODE: Using hardcoded Nifty 50 universe (skipping yfinance loop) …")

    # Current Nifty 50 constituents (NSE symbols).
    # Note: HDFC Ltd merged into HDFCBANK in 2023 and is no longer listed.
    # JIOFIN, ETERNAL (Zomato), BEL, TRENT are recent additions.
    nifty50_data = [
        ("ADANIENT",   "Adani Enterprises Ltd",                      250_000),
        ("ADANIPORTS",  "Adani Ports & SEZ Ltd",                      310_000),
        ("APOLLOHOSP",  "Apollo Hospitals Enterprise Ltd",             100_000),
        ("ASIANPAINT",  "Asian Paints Ltd",                            195_000),
        ("AXISBANK",    "Axis Bank Ltd",                               370_000),
        ("BAJAJ-AUTO",  "Bajaj Auto Ltd",                              225_000),
        ("BAJAJFINSV",  "Bajaj Finserv Ltd",                           270_000),
        ("BAJFINANCE",  "Bajaj Finance Ltd",                           440_000),
        ("BEL",         "Bharat Electronics Ltd",                      200_000),
        ("BHARTIARTL",  "Bharti Airtel Ltd",                           950_000),
        ("BRITANNIA",   "Britannia Industries Ltd",                    115_000),
        ("CIPLA",       "Cipla Ltd",                                   120_000),
        ("COALINDIA",   "Coal India Ltd",                              240_000),
        ("DIVISLAB",    "Divi's Laboratories Ltd",                     130_000),
        ("DRREDDY",     "Dr. Reddy's Laboratories Ltd",                140_000),
        ("EICHERMOT",   "Eicher Motors Ltd",                           130_000),
        ("ETERNAL",     "Eternal Ltd (Zomato)",                        220_000),
        ("GRASIM",      "Grasim Industries Ltd",                       175_000),
        ("HCLTECH",     "HCL Technologies Ltd",                        440_000),
        ("HDFCBANK",    "HDFC Bank Ltd",                             1_450_000),
        ("HDFCLIFE",    "HDFC Life Insurance Company Ltd",             140_000),
        ("HEROMOTOCO",  "Hero MotoCorp Ltd",                           100_000),
        ("HINDALCO",    "Hindalco Industries Ltd",                     200_000),
        ("HINDUNILVR",  "Hindustan Unilever Ltd",                      530_000),
        ("ICICIBANK",   "ICICI Bank Ltd",                              900_000),
        ("INFY",        "Infosys Ltd",                                 680_000),
        ("ITC",         "ITC Ltd",                                     500_000),
        ("JIOFIN",      "Jio Financial Services Ltd",                  195_000),
        ("JSWSTEEL",    "JSW Steel Ltd",                               240_000),
        ("KOTAKBANK",   "Kotak Mahindra Bank Ltd",                     390_000),
        ("LT",          "Larsen & Toubro Ltd",                         500_000),
        ("M&M",         "Mahindra & Mahindra Ltd",                     370_000),
        ("MARUTI",      "Maruti Suzuki India Ltd",                     360_000),
        ("NESTLEIND",   "Nestle India Ltd",                            225_000),
        ("NTPC",        "NTPC Ltd",                                    370_000),
        ("ONGC",        "Oil & Natural Gas Corporation Ltd",           310_000),
        ("POWERGRID",   "Power Grid Corporation of India Ltd",         290_000),
        ("RELIANCE",    "Reliance Industries Ltd",                   1_800_000),
        ("SBILIFE",     "SBI Life Insurance Company Ltd",              145_000),
        ("SBIN",        "State Bank of India",                         700_000),
        ("SUNPHARMA",   "Sun Pharmaceutical Industries Ltd",           390_000),
        ("TATACONSUM",  "Tata Consumer Products Ltd",                  115_000),
        ("TATAMOTORS",  "Tata Motors Ltd",                             310_000),
        ("TATASTEEL",   "Tata Steel Ltd",                              180_000),
        ("TCS",         "Tata Consultancy Services Ltd",             1_400_000),
        ("TECHM",       "Tech Mahindra Ltd",                           165_000),
        ("TITAN",       "Titan Company Ltd",                           275_000),
        ("TRENT",       "Trent Ltd",                                   175_000),
        ("ULTRACEMCO",  "UltraTech Cement Ltd",                        300_000),
        ("WIPRO",       "Wipro Ltd",                                   265_000),
    ]

    df = pd.DataFrame(nifty50_data, columns=["symbol", "company_name", "market_cap_crores"])
    log.info(f"  ✓ Nifty 50 universe loaded: {len(df)} companies")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — MODULAR FILTERING PIPELINE
# ─────────────────────────────────────────────────────────────────────────────
# Each filter function MUST:
#   • Accept a pd.DataFrame  (columns: symbol, company_name, market_cap_crores)
#   • Return a filtered pd.DataFrame
#   • Log how many rows were retained
#
# To add a new filter: write the function, append it to FILTER_PIPELINE list.
# ─────────────────────────────────────────────────────────────────────────────

def filter_by_market_cap(df: pd.DataFrame) -> pd.DataFrame:
    """[ACTIVE] Keep companies with market cap ≥ MIN_MARKET_CAP_CRORES."""
    before = len(df)
    df = df[df["market_cap_crores"] >= MIN_MARKET_CAP_CRORES].copy()
    log.info(f"  filter_by_market_cap  → {len(df)}/{before} companies retained "
             f"(≥ ₹{MIN_MARKET_CAP_CRORES:,} Cr)")
    return df


# ── TEMPLATE FILTERS — uncomment & implement when needed ─────────────────────

# def filter_by_pe_ratio(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep companies with PE ratio ≤ MAX_PE."""
#     MAX_PE = 40
#     # TODO: add 'pe_ratio' column to df before this filter runs
#     before = len(df)
#     df = df[df["pe_ratio"] <= MAX_PE].copy()
#     log.info(f"  filter_by_pe_ratio    → {len(df)}/{before} retained (PE ≤ {MAX_PE})")
#     return df


# def filter_by_sector(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep only companies in ALLOWED_SECTORS."""
#     ALLOWED_SECTORS = {"FMCG", "Pharmaceuticals", "Technology", "Banking"}
#     # TODO: add 'sector' column to df before this filter runs
#     before = len(df)
#     df = df[df["sector"].isin(ALLOWED_SECTORS)].copy()
#     log.info(f"  filter_by_sector      → {len(df)}/{before} retained")
#     return df


# def filter_by_revenue_growth(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep companies with YoY revenue growth ≥ threshold."""
#     MIN_REVENUE_GROWTH_PCT = 10
#     before = len(df)
#     df = df[df["revenue_growth_pct"] >= MIN_REVENUE_GROWTH_PCT].copy()
#     log.info(f"  filter_by_revenue_growth → {len(df)}/{before} retained")
#     return df


# def filter_by_debt_to_equity(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Remove highly leveraged companies."""
#     MAX_DE_RATIO = 1.5
#     before = len(df)
#     df = df[df["de_ratio"] <= MAX_DE_RATIO].copy()
#     log.info(f"  filter_by_de_ratio    → {len(df)}/{before} retained")
#     return df


# def filter_by_promoter_holding(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep companies where promoter holding ≥ threshold."""
#     MIN_PROMOTER_HOLDING_PCT = 30
#     before = len(df)
#     df = df[df["promoter_holding_pct"] >= MIN_PROMOTER_HOLDING_PCT].copy()
#     log.info(f"  filter_by_promoter    → {len(df)}/{before} retained")
#     return df


# ── Register active filters here (order matters) ─────────────────────────────
FILTER_PIPELINE = [
    filter_by_market_cap,
    # filter_by_pe_ratio,
    # filter_by_sector,
    # filter_by_revenue_growth,
    # filter_by_debt_to_equity,
    # filter_by_promoter_holding,
]


def run_filter_pipeline(df: pd.DataFrame) -> pd.DataFrame:
    """Run df through every filter in FILTER_PIPELINE sequentially."""
    log.info(f"Starting filter pipeline with {len(df)} companies …")
    for fn in FILTER_PIPELINE:
        df = fn(df)
        if df.empty:
            log.warning("  Pipeline produced 0 results — check filter thresholds!")
            break
    log.info(f"  → {len(df)} companies passed all filters\n")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — TRANSCRIPT DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────
# DESIGN NOTE ON ANTI-SCRAPING:
#
#   BSE India & Screener.in use Cloudflare + dynamic tokens.  Two safe paths:
#
#   OPTION A (recommended): BSE XBRL open API
#     BSE provides an official XML/JSON filings feed. Concall transcripts
#     are filed under 'Corporate Announcements' (categoryId=6 or subcategory
#     "Conference Call" / "Investor Presentation"). No auth needed; just
#     rotate User-Agents and respect rate limits.
#
#   OPTION B: Screener.in unofficial wrapper
#     Library: https://github.com/pratik2315/screener-python (community)
#     Install: pip install screener-python
#     Caveat: token-based, may break on UI changes — use as fallback only.
#
#   OPTION C: NSE XBRL feed (alternative to BSE)
#     NSE also publishes announcements via:
#     https://www.nseindia.com/api/corporate-announcements
# ─────────────────────────────────────────────────────────────────────────────

# Realistic browser User-Agent pool for rotation
_USER_AGENTS = [
    ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
     "AppleWebKit/537.36 (KHTML, like Gecko) "
     "Chrome/124.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) "
     "AppleWebKit/605.1.15 (KHTML, like Gecko) "
     "Version/17.4 Safari/605.1.15"),
    ("Mozilla/5.0 (X11; Linux x86_64; rv:125.0) "
     "Gecko/20100101 Firefox/125.0"),
]


def _get_headers(referer: str = "https://www.bseindia.com") -> dict:
    """Return headers that mimic a real browser, with rotated User-Agent."""
    return {
        "User-Agent":      random.choice(_USER_AGENTS),
        "Accept":          "application/json, text/plain, */*",
        "Accept-Language": "en-IN,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer":         referer,
        "DNT":             "1",
        "Connection":      "keep-alive",
    }


def _polite_sleep():
    """Sleep a random interval to avoid getting blocked."""
    secs = random.uniform(SLEEP_MIN_SECS, SLEEP_MAX_SECS)
    time.sleep(secs)


import json
import time
import logging
import requests
import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

log = logging.getLogger("ConcallAnalyzer")

class ScreenerTranscriptFetcher:
    def __init__(self):
        log.info("  Booting up stealth browser engine for Screener...")
        self.playwright = sync_playwright().start()
        
        # NOTE: If you still get blocked, change headless=True to headless=False 
        # so you can see the browser and manually click any Cloudflare boxes!
        self.browser = self.playwright.chromium.launch(headless=False) 
        self.context = self.browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        self.page = self.context.new_page()
        self.nse_page = self.context.new_page()

        # Warm up the browser on Screener's homepage to grab initial cookies
        try:
            self.page.goto("https://www.screener.in", timeout=60000)
            self.page.wait_for_timeout(3000)
        except Exception as e:
            log.debug(f"  Browser warmup timeout: {e}")

    def __del__(self):
        # Clean up the browser when the script finishes
        try:
            if hasattr(self, 'browser'):
                self.browser.close()
            if hasattr(self, 'playwright'):
                self.playwright.stop()
        except Exception:
            pass

    def get_bse_code(self, nse_symbol: str):
        # We don't need BSE codes for Screener! Just return the NSE symbol.
        return nse_symbol.strip().upper()

    def fetch_latest_transcript_url(self, symbol: str, lookback_days: int = 180):
        url = f"https://www.screener.in/company/{symbol}/consolidated/"
        try:
            self.page.goto(url, timeout=30000)
            self.page.wait_for_timeout(2000) # Give the page time to load
            
            # Check if we hit a security wall
            page_title = self.page.title()
            if "Just a moment" in page_title or "Cloudflare" in page_title:
                log.warning(f"  [!] Security Check triggered for {symbol}! Please click the checkbox in the browser...")
                self.page.wait_for_timeout(10000) # Wait 10 seconds for you to click it
            
            html = self.page.content()
            soup = BeautifulSoup(html, 'html.parser')
            
            # FOOLPROOF SEARCH: Scan EVERY link on the page for "Transcript" or "Concall"
            links = soup.find_all('a', string=lambda text: text and ('Transcript' in text or 'Concall' in text))
            
            if links:
                # Grab the very first link that matches
                pdf_link = links[0].get('href')
                
                if pdf_link.startswith('/'):
                    pdf_link = "https://www.screener.in" + pdf_link
                    
                return pdf_link
                    
        except Exception as e:
            log.debug(f"  Screener browser fetch error for {symbol}: {e}")
            
        return None

    def fetch_transcript_url_from_nse(self, symbol: str, lookback_days: int = 180):
        """
        Fallback: find concall transcript on NSE using a separate browser tab.
        Uses self.nse_page so the Screener session on self.page stays intact.
        """
        try:
            from datetime import datetime, timedelta

            # Warm up NSE on the separate tab to get session cookies
            self.nse_page.goto("https://www.nseindia.com", timeout=30000)
            self.nse_page.wait_for_timeout(2000)

            api_url = (
                f"https://www.nseindia.com/api/corporate-announcements"
                f"?index=equities&symbol={symbol}"
            )

            response = self.nse_page.evaluate(f"""
                async () => {{
                    const r = await fetch("{api_url}", {{
                        headers: {{
                            "Accept": "application/json",
                            "Referer": "https://www.nseindia.com/"
                        }}
                    }});
                    return await r.json();
                }}
            """)

            if not response or not isinstance(response, list):
                log.warning(f"  [{symbol}] NSE returned no data: {response}")
                return None
            
            
            cutoff = datetime.today() - timedelta(days=lookback_days)

            for item in response:
                
                desc = (item.get("desc", "") or "").lower()
                an_dt_str = item.get("an_dt", "")

                # Must mention transcript/concall
                if not any(kw in desc for kw in ("concall", "transcript", "conference call", "earnings call", "con. call")):
                    continue
                # Skip schedule notices — these are invitations, not actual transcripts
                attchmnt_text = (item.get("attchmntText", "") or "").lower()
                if any(kw in attchmnt_text for kw in ("schedule of meet", "schedule of analyst", "intimation of meet", "inform the exchange about schedule")):
                    continue
                try:
                    an_dt = datetime.strptime(an_dt_str.split(" ")[0], "%d-%b-%Y")
                    if an_dt < cutoff:
                        continue
                except Exception:
                    # If date can't be parsed, skip this item to be safe
                    continue

                pdf_file = item.get("attchmntFile", "")
                if pdf_file:
                    # attchmntFile sometimes returns a full URL, sometimes just a filename
                    if pdf_file.startswith("http"):
                        pdf_url = pdf_file
                    else:
                        pdf_url = f"https://nsearchives.nseindia.com/corporate/{pdf_file}"
                    log.info(f"  [{symbol}] Found on NSE: {pdf_file}")
                    return pdf_url

        except Exception as e:
            log.debug(f"  [{symbol}] NSE fallback error: {e}")

        return None

    def download_pdf(self, pdf_url: str, save_path: Path):
        """
        Download the PDF using stealth headers to bypass BSE/Screener blocks.
        """
        # THE FIX: Tell the server we are a real Chrome browser coming from BSE
        stealth_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/pdf,application/xhtml+xml,text/html;q=0.9,*/*;q=0.8",
            "Referer": "https://www.bseindia.com/"  # Critical for BSE links!
        }
        
        try:
            # Pass the headers into the request
            response = requests.get(pdf_url, headers=stealth_headers, timeout=15)
            response.raise_for_status()
            
            # Check if we accidentally got an HTML page anyway
            if b"<!DOC" in response.content[:10].upper() or b"<HTML" in response.content[:10].upper():
                log.warning("Server blocked the PDF and sent an HTML page instead.")
                return False

            # Save the real PDF
            with open(save_path, "wb") as f:
                f.write(response.content)
            return True
            
        except Exception as e:
            log.error(f"Failed to download PDF: {e}")
            return False
# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — PDF TEXT PARSING
# ─────────────────────────────────────────────────────────────────────────────

def extract_text_from_pdf(pdf_path: Path) -> str:
    """
    Extract raw text from a PDF with validation and dual-parser fallback.
    """
    # ── Layer 1: Validate ────────────────────────────────────────────────────
    try:
        file_size = pdf_path.stat().st_size
        if file_size < 5_120:
            log.warning(f"  [{pdf_path.stem}] Too small ({file_size} bytes) — deleting cache.")
            pdf_path.unlink(missing_ok=True)
            return ""
        with open(pdf_path, "rb") as f:
            raw = f.read(1024)
        header_check = raw.lstrip()[:5]
        if header_check != b"%PDF-":
            log.warning(f"  [{pdf_path.stem}] Not a valid PDF (got {header_check}) — deleting cache.")
            pdf_path.unlink(missing_ok=True)
            return ""
    except Exception as e:
        log.warning(f"  [{pdf_path.stem}] Validation error: {e}")
        return ""

    # ── Layer 2: pdfplumber ──────────────────────────────────────────────────
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            pages_text = []
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    pages_text.append(t)
            text = "\n".join(pages_text)
        if text.strip():
            return text
    except ImportError:
        pass
    except Exception as e:
        log.debug(f"  [{pdf_path.stem}] pdfplumber error: {e}")

    # ── Layer 3: pypdf ───────────────────────────────────────────────────────
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        pages_text = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                pages_text.append(t)
        text = "\n".join(pages_text)
        if text.strip():
            return text
    except ImportError:
        log.error("  pypdf not installed. Run: pip install pypdf")
    except Exception as e:
        log.debug(f"  [{pdf_path.stem}] pypdf error: {e}")

    return ""

def clean_text(text: str) -> str:
    """
    Normalise transcript text for keyword scanning:
      - Lowercase
      - Collapse whitespace / newlines
      - Remove non-alphanumeric characters except spaces and hyphens
    """
    text = text.lower()
    text = re.sub(r"[\r\n\t]+", " ", text)          # flatten newlines
    text = re.sub(r"[^a-z0-9 \-]", " ", text)       # strip punctuation
    text = re.sub(r"\s{2,}", " ", text).strip()      # collapse spaces
    return text


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — KEYWORD SCORING
# ─────────────────────────────────────────────────────────────────────────────

# Layer 1 — minimum % threshold for the Revenue Growth Rule
# Set to 10.0 for the Nifty 50 test run (large-caps grow slower; 20% would miss most).
# Change back to 20.0 for the full NSE universe run.
QUANTITATIVE_MIN_PCT: float = 20.0

# Layer 2 — minimum cosine similarity score to count as a semantic match
SEMANTIC_SIMILARITY_THRESHOLD: float = 0.75

# Layer 3 — DistilBERT negative-score threshold (0–1 scale).
# Sentences with a NEGATIVE score above this are discarded.
SENTIMENT_NEGATIVE_THRESHOLD: float = 0.60

# Maximum number of quote snippets written to context_snippets (Excel cell limit)
MAX_SNIPPETS: int = 7

# Sentence minimum length filter (characters)
MIN_SENTENCE_LEN: int = 25


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 1 ─ QUANTITATIVE EXTRACTION PATTERNS
# ─────────────────────────────────────────────────────────────────────────────
# Each category maps to a list of regex patterns.  Every pattern MUST contain
# exactly one capture group that extracts the percentage number.
# Patterns are tried in order; the first match wins.
# ─────────────────────────────────────────────────────────────────────────────

QUANTITATIVE_PATTERNS: dict[str, list[str]] = {

    "High Growth / Topline": [
        # "revenue/sales grew/increased/jumped by 25%"
        r"(?:topline|top[\-\s]line|revenue|net\s+sales?|sales?|turnover)"
        r"[\w\s,;]+?(?:grew?|grow(?:ing|th)|increas(?:ed?|ing)|jump(?:ed)?|"
        r"surged?|rose?|up|higher)\s+(?:by\s+)?(\d+(?:\.\d+)?)\s*%",
        # "revenue growth of 25%" / "sales growth of 8%" — noun phrase form
        # MUST be here so Layer 1 intercepts it and enforces the 20% threshold
        # before it falls through to the keyword fallback which has no % check.
        r"(?:topline|top[\-\s]line|revenue|net\s+sales?|sales?|turnover)"
        r"\s+growth\s+of\s+(\d+(?:\.\d+)?)\s*%",
        # "25% revenue growth / 25% growth in revenue"
        r"(\d+(?:\.\d+)?)\s*%\s+(?:revenue|sales?|topline|top[\-\s]line|"
        r"turnover)\s+(?:growth|increase|jump|surge)",
        r"(\d+(?:\.\d+)?)\s*%\s+(?:growth|increase|jump|surge)\s+in\s+"
        r"(?:revenue|sales?|topline)",
        # "achieved/delivered/posted 30% revenue growth"
        r"(?:achiev|record|report|deliver|post|clock)(?:ed?|ing)\s+"
        r"(\d+(?:\.\d+)?)\s*%\s+(?:revenue|sales?|topline|top[\-\s]line)",
        # "22% yoy revenue growth"
        r"(\d+(?:\.\d+)?)\s*%\s+(?:yoy|y[\-\s]o[\-\s]y|year[\-\s]on[\-\s]year|"
        r"qoq|q[\-\s]o[\-\s]q|quarter[\-\s]on[\-\s]quarter)\s+"
        r"(?:revenue|sales?|growth|topline)",
    ],

    "Order Visibility": [
        # "order inflows of INR X crores" — no % threshold applied to this category
        # because order book sizes are typically in absolute INR, not %.
        # We keep it in QUANTITATIVE_PATTERNS as an optional trigger.
        # Use a dummy pattern that intentionally never fires (order book
        # matching is handled fully by the keyword layer below).
        r"order\s+(?:book|inflow|backlog)\s+(?:at|of|stands?\s+at)\s+"
        r"(?:inr|rs\.?|₹)?\s*(\d+(?:\.\d+)?)\s*(?:crore|cr|lakh|mn|bn)",
    ],

    "Margin Expansion": [
        # "EBITDA margins expanded/improved by 200 bps to 22%"
        r"(?:ebitda|ebit|gross|operating|net|pbt)\s+margins?\s+"
        r"[\w\s,;]+?(?:expand|improv|increas|up|higher|widen|grew?)\s+"
        r"(?:by\s+)?(\d+(?:\.\d+)?)\s*%",
        # "margins expanded 150 bps" (basis points → treated as pct for threshold)
        r"margins?\s+[\w\s]+?(?:expand|improv|up|widen)\s+(?:by\s+)?(\d+)\s*bps",
        # "margin now at 25%" — absolute margin level used as proxy
        r"(?:ebitda|ebit|gross|operating)\s+margins?\s+(?:of|at|to)\s+"
        r"(\d+(?:\.\d+)?)\s*%",
    ],

    "Return Ratios": [
        # "ROCE improved to 22%"
        r"(?:roce|roe|roic|roi|return\s+on\s+(?:capital|equity|invest))"
        r"[\w\s,;]+?(?:improv|increas|expand|up|higher|rose?|grew?)\s+"
        r"(?:by\s+)?(?:to\s+)?(\d+(?:\.\d+)?)\s*%",
        r"(?:roce|roe|roic|roi)\s+(?:of|at|now|stands?\s+at)\s+"
        r"(\d+(?:\.\d+)?)\s*%",
    ],

    "Market Share": [
        # "market share grew from 12% to 17%"
        r"market\s+share\s+[\w\s,;]+?(?:grew?|up|higher|from\s+\d+%?\s+to)\s+"
        r"(?:to\s+)?(\d+(?:\.\d+)?)\s*%",
        r"(?:market|wallet)\s+share\s+(?:of|at|to)\s+(\d+(?:\.\d+)?)\s*%",
    ],
}


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 2 ─ SEMANTIC SEARCH CONCEPT LIBRARY
# ─────────────────────────────────────────────────────────────────────────────
# These are the "concept anchors" that will be encoded into embedding vectors.
# Add richer paraphrase variations to improve recall.
# ─────────────────────────────────────────────────────────────────────────────

SEMANTIC_CONCEPTS: dict[str, list[str]] = {

    "Margin Expansion": [
        "EBITDA margin expansion and profitability improvement",
        "gross margin improved due to better product mix",
        "operating leverage driving margin expansion",
        "margins expanded as fixed costs got absorbed on higher revenues",
        "cost efficiencies leading to margin improvement",
        "better realisations improving profitability",
        "premium product mix driving margin accretion",
    ],

    "Capacity Expansion": [
        "new greenfield plant commissioned and ramping up",
        "brownfield capacity expansion adding production capacity",
        "capex investment in new manufacturing facility",
        "capacity addition to meet growing demand",
        "new plant going on stream increasing installed capacity",
        "expanding production footprint to scale operations",
    ],

    "Deleveraging": [
        "significant debt repayment reducing leverage on balance sheet",
        "net debt to EBITDA ratio coming down",
        "balance sheet deleveraging with debt reduction",
        "company becoming debt free after debt repayment",
        "leverage ratio improved after repaying term loans",
        "strong free cash flows used for debt reduction",
    ],

    "Integration & Efficiency": [
        "backward integration reducing raw material dependency",
        "forward integration into higher value-added products",
        "debottlenecking increasing throughput without capex",
        "economies of scale improving operational efficiency",
        "process automation driving cost savings",
        "vertical integration improving margin and supply security",
    ],

    "Cash Generation": [
        "strong free cash flow generation supporting growth investments",
        "company turned net cash positive with no net debt",
        "FCF positive and self-funding growth capex",
        "robust operating cash flows funding capex and debt repayment",
        "net debt negative balance sheet with cash surplus",
    ],

    "Working Capital": [
        "working capital days reduced improving cash conversion",
        "inventory days and receivable days improved",
        "cash conversion cycle shortened due to better collections",
        "debtors days reduced as collection efficiency improved",
        "working capital management improving operating cash flow",
    ],

    "Order Visibility": [
        "strong order book providing revenue visibility",
        "new order inflows at record high levels",
        "robust pipeline of orders from key customers",
        "order backlog covering multiple quarters of revenue",
        "large order wins strengthening growth visibility",
    ],

    "Corporate Action": [
        "strategic acquisition adding capabilities and market access",
        "merger creating synergies and scale benefits",
        "demerger unlocking value for shareholders",
        "corporate restructuring simplifying group structure",
        "acquisition of complementary business expanding addressable market",
    ],
}


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 3 ─ NEGATIVE SENTIMENT SIGNAL LIBRARY (regex fallback)
# ─────────────────────────────────────────────────────────────────────────────

# These patterns are used ONLY when the DistilBERT model cannot be loaded.
_NEG_SIGNAL_PATTERNS: list[str] = [
    r"\b(?:delay(?:ed?|s|ing)?|postpone(?:d|s|ing)?|setback)\b",
    r"\b(?:declin(?:ed?|ing|e)?|decreas(?:ed?|ing|e)?|fell?|fall(?:ing)?|drop(?:ped?|ping)?)\b",
    r"\b(?:challeng(?:ed?|ing|es?)|difficult(?:y|ies)?|headwind[s]?|pressure[sd]?)\b",
    r"\b(?:below\s+(?:expectation|guidance|target)|miss(?:ed?|ing)?|shortfall|underperform)\b",
    r"\b(?:took\s+a\s+hit|margin\s+compression|erosion|adversely|negatively\s+impact)\b",
    r"\b(?:not\s+able|unable|couldn.t|could\s+not|failed?\s+to|did\s+not\s+achiev)\b",
    r"\b(?:concerns?|uncertainty|uncertainties|cautious|subdued|muted)\b",
    r"\b(?:slowdown|weakness|weak(?:ened?)?|soft(?:ness)?|sluggish)\b",
    r"\b(?:elevated\s+(?:cost[s]?|rm|input|raw\s+material))\b",
    r"\b(?:impact(?:ed)?\s+by|affected\s+by|hurt\s+by|weighed\s+by|dragged\s+by)\b",
]

# These patterns alone are enough to discard a sentence regardless of other signals
_STRONG_NEG_OVERRIDES: list[str] = [
    r"\b(?:delay(?:ed?|s|ing)?|postpone(?:d|s|ing)?)\b",
    r"\btook\s+a\s+hit\b",
    r"\bheadwind[s]?\b",
    r"\belevated\s+(?:cost[s]?|rm|input|raw\s+material)\b",
    r"\bbelow\s+(?:expectation[s]?|guidance)\b",
    r"\bmargin\s+compression\b",
    r"\badversely\s+impact\b",
    r"\b(?:halt(?:ed|s|ing)?|pause[sd]?|stop(?:ped)?)\b",
]

# Pre-compile for speed
_NEG_COMPILED = [re.compile(p) for p in _NEG_SIGNAL_PATTERNS]
_STRONG_NEG_COMPILED = [re.compile(p) for p in _STRONG_NEG_OVERRIDES]


# ─────────────────────────────────────────────────────────────────────────────
# MODEL INITIALISATION  ─ call nlp_engine_init() ONCE at startup
# ─────────────────────────────────────────────────────────────────────────────

# Module-level singletons so models are loaded only once per process
_spacy_nlp = None           # spaCy pipeline (sentence tokenizer)
_semantic_model = None      # SentenceTransformer
_concept_embeddings = None  # Pre-computed embeddings for SEMANTIC_CONCEPTS
_sentiment_pipeline = None  # HuggingFace sentiment classifier
_use_regex_sentiment = False  # Fallback flag


def nlp_engine_init() -> None:
    """
    Load all ML models into module-level singletons.
    Call this ONCE at the top of main() or demo_mode() before the scoring loop.

    Models downloaded on first run (~80 MB for MiniLM + ~250 MB for DistilBERT).
    Subsequent runs load from local cache — no internet required.
    """
    global _spacy_nlp, _semantic_model, _concept_embeddings
    global _sentiment_pipeline, _use_regex_sentiment

    # ── spaCy sentence tokeniser ─────────────────────────────────────────────
    log.info("NLP Engine: loading spaCy en_core_web_sm …")
    try:
        import spacy
        _spacy_nlp = spacy.load("en_core_web_sm", disable=["ner", "tagger", "lemmatizer"])
        # Increase max length for long transcripts
        _spacy_nlp.max_length = 2_000_000
        log.info("  ✓ spaCy loaded")
    except Exception as e:
        log.warning(f"  spaCy not available ({e}). Using fallback sentence splitter.")
        _spacy_nlp = None

    # ── SentenceTransformers (Layer 2 — semantic search) ─────────────────────
    log.info("NLP Engine: loading SentenceTransformer all-MiniLM-L6-v2 …")
    try:
        from sentence_transformers import SentenceTransformer
        _semantic_model = SentenceTransformer("all-MiniLM-L6-v2")
        # Pre-compute embeddings for all concept phrases (done once, reused per transcript)
        all_concepts = []
        concept_category_index = []
        for cat, phrases in SEMANTIC_CONCEPTS.items():
            for phrase in phrases:
                all_concepts.append(phrase)
                concept_category_index.append(cat)
        concept_vecs = _semantic_model.encode(all_concepts, normalize_embeddings=True,
                                              show_progress_bar=False)
        _concept_embeddings = (concept_vecs, concept_category_index)
        log.info(f"  ✓ SentenceTransformer loaded | {len(all_concepts)} concept anchors encoded")
    except Exception as e:
        log.warning(f"  SentenceTransformer not available ({e}). Semantic search disabled.")
        _semantic_model = None
        _concept_embeddings = None

    # ── HuggingFace Sentiment (Layer 3 — tone detection) ─────────────────────
    log.info("NLP Engine: loading DistilBERT sentiment classifier …")
    try:
        from transformers import pipeline as hf_pipeline
        _sentiment_pipeline = hf_pipeline(
            "sentiment-analysis",
            model="distilbert-base-uncased-finetuned-sst-2-english",
            device=-1,          # CPU only; change to 0 for GPU if available
            truncation=True,
            max_length=128,
        )
        _use_regex_sentiment = False
        log.info("  ✓ DistilBERT sentiment classifier loaded")
    except Exception as e:
        log.warning(f"  DistilBERT not available ({e}). Using regex sentiment fallback.")
        _sentiment_pipeline = None
        _use_regex_sentiment = True

    log.info("NLP Engine: initialisation complete\n")


# ─────────────────────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _split_sentences(text: str) -> list[str]:
    """
    Split transcript text into individual sentences.
    Uses spaCy if loaded, otherwise falls back to a reliable regex splitter.
    """
    if _spacy_nlp is not None:
        try:
            doc = _spacy_nlp(text)
            return [sent.text.strip() for sent in doc.sents
                    if len(sent.text.strip()) >= MIN_SENTENCE_LEN]
        except Exception:
            pass  # fall through to regex splitter

    # Regex fallback: split on ". ", "! ", "? " but keep common abbreviations intact
    text_flat = re.sub(r"[\r\n\t]+", " ", text)
    # Protect common abbreviations
    text_flat = re.sub(r"\b(Mr|Mrs|Ms|Dr|vs|etc|approx|Co|Ltd|Pvt|INR|Rs|FY|Q[1-4])\.", r"\1<DOT>", text_flat)
    # Split
    raw_sents = re.split(r"(?<=[.!?])\s+", text_flat)
    # Restore dots
    sents = [s.replace("<DOT>", ".").strip() for s in raw_sents]
    return [s for s in sents if len(s) >= MIN_SENTENCE_LEN]


def _extract_percentage_from_match(m: re.Match) -> Optional[float]:
    """Safely extract and return the captured percentage from a regex match."""
    try:
        return float(m.group(1))
    except (IndexError, ValueError, TypeError):
        return None


def _layer1_quantitative(sentence: str) -> tuple[Optional[str], Optional[float], Optional[str]]:
    """
    Layer 1: Scan a sentence for a metric-attached percentage ≥ QUANTITATIVE_MIN_PCT.

    Returns:
        (category, percentage, trigger_label) if found, else (None, None, None)
    """
    s = sentence.lower()
    # Normalise number formats: "25%" / "25 %" / "25 per cent" / "25 percent"
    s = re.sub(r"(\d+(?:\.\d+)?)\s*per\s*cent", r"\1%", s)

    for category, patterns in QUANTITATIVE_PATTERNS.items():
        for pat in patterns:
            m = re.search(pat, s)
            if m:
                pct = _extract_percentage_from_match(m)
                if pct is None:
                    continue
                
                # BPS logic remains the same
                if "bps" in pat:
                    return (category, pct, f"{int(pct)} bps")
                
                # The 20% Rule ONLY applies to Topline Growth
                if category == "High Growth / Topline":
                    if pct >= QUANTITATIVE_MIN_PCT:
                        return (category, pct, f"Growth ≥ {pct:.0f}%")
                    else:
                        continue # Silently drop it if revenue growth is under 20%
                else:
                    # For Margins, Market Share, etc., accept any extracted percentage
                    return (category, pct, f"{pct:.0f}%")
    return (None, None, None)


def _layer2_semantic(sentence: str) -> Optional[str]:
    """
    Layer 2: Compute cosine similarity between the sentence embedding and
    pre-computed concept embeddings.  Return the best-matching category
    if similarity exceeds SEMANTIC_SIMILARITY_THRESHOLD, else None.

    Returns None if SentenceTransformers is not initialised.
    """
    if _semantic_model is None or _concept_embeddings is None:
        return None

    concept_vecs, concept_category_index = _concept_embeddings

    try:
        sent_vec = _semantic_model.encode([sentence], normalize_embeddings=True,
                                          show_progress_bar=False)[0]
        # Cosine similarity = dot product of unit vectors
        sims = concept_vecs @ sent_vec
        best_idx = int(np.argmax(sims))
        best_score = float(sims[best_idx])
        if best_score >= SEMANTIC_SIMILARITY_THRESHOLD:
            return concept_category_index[best_idx]
    except Exception:
        pass

    return None


def _layer3_is_negative(sentence: str) -> bool:
    """
    Layer 3: Return True if the sentence has negative/headwind tone.

    Special case: "despite headwinds / notwithstanding challenges" is POSITIVE
    framing — the company overcame the obstacle. These are checked first and
    short-circuit the negative detection entirely.
    """
    s = sentence.lower()

    # ── 0. CONCESSION GUARD ───────────────────────────────────────────────────
    # "Revenue grew 22% despite headwinds" is positive — the company overcame it.
    # If a concession word immediately precedes or follows a negative word,
    # treat the whole sentence as positive and skip all further checks.
    _DESPITE_PAT = re.compile(
        r"\b(?:despite|notwithstanding|in\s+spite\s+of)\b[\w\s,;]{0,60}"
        r"(?:headwind|challenge|difficult|pressure|setback|weak|adverse|macro)",
        re.I,
    )
    _DESPITE_REV_PAT = re.compile(
        r"\b(?:headwind|challenge|difficult|pressure|setback|weak|adverse|macro)"
        r"[\w\s,;]{0,60}\b(?:despite|notwithstanding|in\s+spite\s+of)\b",
        re.I,
    )
    if _DESPITE_PAT.search(sentence) or _DESPITE_REV_PAT.search(sentence):
        return False  # positive framing — keep this sentence

    # ── 1. ABSOLUTE DEALBREAKERS (Runs First) ─────────────────────────────────
    # If the CEO mentions a delay, halt, or hit, we instantly throw it in the 
    # trash, even if they try to spin it as an "opportunity".
    if any(p.search(s) for p in _STRONG_NEG_COMPILED):
        return True

    # ── 2. DistilBERT AI Path (For subtler negativity) ────────────────────────
    if _sentiment_pipeline is not None and not _use_regex_sentiment:
        try:
            result = _sentiment_pipeline(sentence[:512])[0]
            if result["label"] == "NEGATIVE" and result["score"] >= SENTIMENT_NEGATIVE_THRESHOLD:
                return True
            return False
        except Exception:
            pass  # fall through to regex fallback

    # ── 3. Regex Fallback (For minor soft signals) ────────────────────────────
    soft_hits = sum(1 for p in _NEG_COMPILED if p.search(s))
    return soft_hits >= 2

def _extract_all_percentages(sentence: str) -> list[float]:
    """Return every percentage number in a sentence (handles 'per cent' spelling too)."""
    s = re.sub(r"(\d+(?:\.\d+)?)\s*per\s*cent", r"\1%", sentence.lower())
    return [float(x) for x in re.findall(r"(\d+(?:\.\d+)?)\s*%", s)]


def _also_keyword_matched(sentence: str) -> tuple[Optional[str], Optional[str]]:
    """
    Run the original KEYWORDS_DICT check on a single sentence.
    Returns (category, trigger_word) if matched, else (None, None).

    20% GATE: For 'High Growth / Topline', if the sentence explicitly states
    a percentage AND that percentage is below QUANTITATIVE_MIN_PCT, the match
    is discarded. This closes the leak where sub-20% revenue sentences bypass
    Layer 1 and get scored via the keyword fallback with no threshold check.

    Rule:
      - "revenue growth of 12%"  → keyword matches, 12 < 20 → DISCARDED
      - "revenue growth of 25%"  → keyword matches, 25 >= 20 → KEPT
      - "revenue growth momentum" → keyword matches, no % stated → KEPT
        (purely qualitative claim — no number to fail the threshold)
    """
    s_lower = sentence.lower()
    search_text = re.sub(r"[^a-z0-9 \-]", " ", s_lower)
    for category, synonyms in KEYWORDS_DICT.items():
        for synonym in synonyms:
            pat = re.compile(r"(?<![a-z])" + re.escape(synonym) + r"(?![a-z])")
            if pat.search(search_text):
                # 20% gate: only applies to the revenue growth category
                if category == "High Growth / Topline":
                    pcts = _extract_all_percentages(sentence)
                    if pcts and max(pcts) < QUANTITATIVE_MIN_PCT:
                        # A % is explicitly stated and it's below threshold — discard
                        return (None, None)
                    # No % stated (qualitative) or max % >= threshold → allow
                return (category, synonym)
    return (None, None)


# ─────────────────────────────────────────────────────────────────────────────
# PRIMARY PUBLIC FUNCTION  —  drop-in replacement for score_transcript()
# ─────────────────────────────────────────────────────────────────────────────

def score_transcript(raw_text: str) -> dict:
    """
    Three-layer NLP scoring engine.  Drop-in replacement for the original
    regex-only score_transcript() — returns the exact same dict structure.

    Returns
    -------
    dict with keys:
        total_score      (int)   — number of distinct categories matched
        categories_found (list)  — list of matched category names
        words_triggered  (list)  — list of trigger phrases / match labels
        priority         (str)   — "High" / "Medium" / "Low" / "Filtered Out"
        context_snippets (str)   — bullet-point formatted matched sentences

    How it works (per sentence)
    ---------------------------
    1. LAYER 1 — Quantitative check: does the sentence contain a recognised
       metric with a number ≥ QUANTITATIVE_MIN_PCT attached?
       → If YES, provisionally flag the sentence with the category.

    2. LAYER 2 — Semantic check (only for sentences that DIDN'T match Layer 1):
       compute cosine similarity to concept anchors.
       → If similarity ≥ SEMANTIC_SIMILARITY_THRESHOLD, provisionally flag.

    3. KEYWORD fallback: for any sentence not caught by Layer 1 or 2, apply
       the original KEYWORDS_DICT regex matching for full backward coverage.

    4. LAYER 3 — Sentiment gate: any provisionally flagged sentence is
       DISCARDED if it carries negative/headwind tone.

    5. Deduplicate categories and build the output dict.
    """
    # ── Pre-process text ──────────────────────────────────────────────────────
    # Preserve punctuation for sentence splitting (do not strip dots)
    text_clean = re.sub(r"[\r\n\t]+", " ", raw_text)
    text_clean = re.sub(r"\s{2,}", " ", text_clean).strip()

    sentences = _split_sentences(text_clean)
    log.debug(f"  Scoring {len(sentences)} sentences …")

    categories_found: set[str] = set()
    words_triggered:  set[str] = set()
    extracted_quotes: list[str] = []

    # ── Per-sentence scoring loop ─────────────────────────────────────────────
    for sentence in sentences:
        provisional_category: Optional[str] = None
        trigger_label: Optional[str] = None
        match_source: str = ""

        # ── LAYER 1: Quantitative ─────────────────────────────────────────────
        cat_l1, pct_l1, label_l1 = _layer1_quantitative(sentence)
        if cat_l1 is not None:
            provisional_category = cat_l1
            trigger_label = f"Exact Metric: {label_l1}"
            match_source = "Metric Match"

        # ── LAYER 2: Semantic (only if Layer 1 did not match) ─────────────────
        if provisional_category is None:
            cat_l2 = _layer2_semantic(sentence)
            if cat_l2 is not None:
                # 20% gate: if the semantic match landed on the revenue category
                # AND the sentence explicitly states a % below the threshold, discard.
                if cat_l2 == "High Growth / Topline":
                    pcts = _extract_all_percentages(sentence)
                    if pcts and max(pcts) < QUANTITATIVE_MIN_PCT:
                        cat_l2 = None  # stated % is too low — reject this match
                if cat_l2 is not None:
                    provisional_category = cat_l2
                    trigger_label = "Conceptual Theme"
                    match_source = "Contextual Match"

        # ── KEYWORD FALLBACK (Now completely merged into Contextual Match) ────
        if provisional_category is None:
            cat_kw, word_kw = _also_keyword_matched(sentence)
            if cat_kw is not None:
                provisional_category = cat_kw
                # We label it as a Contextual Match and show the triggered keyword cleanly
                trigger_label = word_kw
                match_source = "Contextual Match"

        # Nothing matched — move on
        if provisional_category is None:
            continue

        # ── LAYER 3: Sentiment gate ───────────────────────────────────────────
        if _layer3_is_negative(sentence):
            log.debug(f"  [DISCARDED — NEGATIVE TONE] {sentence[:80]}")
            continue

        # ── Commit the match ──────────────────────────────────────────────────
        categories_found.add(provisional_category)
        if trigger_label:
            words_triggered.add(trigger_label)

        quote = f"• {sentence.strip()}"
        if not quote.endswith("."):
            quote += "."
        extracted_quotes.append(quote)

    # ── Scoring & priority ────────────────────────────────────────────────────
    total = len(categories_found)

    if total >= HIGH_PRIORITY_THRESHOLD:
        priority = "High"
    elif total >= MEDIUM_PRIORITY_THRESHOLD:
        priority = "Medium"
    elif total >= LOW_PRIORITY_THRESHOLD:
        priority = "Low"
    else:
        priority = "Filtered Out"

    final_quotes = (
        "\n\n".join(extracted_quotes[:MAX_SNIPPETS])
        if extracted_quotes
        else "No positive catalyst sentences found."
    )

    return {
        "total_score":      total,
        "categories_found": list(categories_found),
        "words_triggered":  list(words_triggered),
        "priority":         priority,
        "context_snippets": final_quotes,
    }



# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — CSV EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_results_to_excel(results: list[dict], discarded: list[dict], output_path: Path):
    """
    Write analysis results and discarded companies to a multi-tab Excel file
    with clickable hyperlinks, color-coding, auto-fitting widths, and text wrapping.
    """
    if not results and not discarded:
        log.warning("No results to export.")
        return

    # Force the file extension to be .xlsx
    excel_path = output_path.with_suffix('.xlsx')
    
    # Prepare the DataFrames
    df_passed = pd.DataFrame([{
        "Ticker":                    row["symbol"],
        "Company Name":              row["company_name"],
        "Market Cap (Crores)":       round(row["market_cap_crores"], 2) if row["market_cap_crores"] else "N/A",
        "Total Categories Matched":  row["total_score"],
        "Categories Found":          ", ".join(row["categories_found"]),
        "Specific Words Triggered":  ", ".join(row["words_triggered"]),
        "Priority Status":           row["priority"],
        "Key Management Quotes":     row["context_snippets"],
        "Transcript Link":           row.get("transcript_url", "N/A"),
    } for row in results])

    df_discarded = pd.DataFrame([{
        "Ticker":                    row["symbol"],
        "Company Name":              row["company_name"],
        "Reason for Discard":        row["reason"]
    } for row in discarded])

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Standard Excel formatting variables
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_text = Font(color="006100")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_text = Font(color="9C5700")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_text = Font(color="9C0006")
    
    # Text Alignment: Top-aligned and Wrapped
    wrap_top_align = Alignment(vertical='top', wrap_text=True)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        
        # --- TAB 1: Qualified Companies ---
        if not df_passed.empty:
            df_passed.to_excel(writer, sheet_name="Qualified Companies", index=False)
            worksheet = writer.sheets["Qualified Companies"]
            
            link_col_idx = df_passed.columns.get_loc("Transcript Link") + 1
            priority_col_idx = df_passed.columns.get_loc("Priority Status") + 1
            
            # Loop 1: Apply Hyperlinks and Color Coding
            for row_num in range(2, worksheet.max_row + 1):
                # 1. Format Hyperlink
                link_cell = worksheet.cell(row=row_num, column=link_col_idx)
                url = link_cell.value
                if url and str(url).startswith("http"):
                    link_cell.hyperlink = url
                    link_cell.value = "View PDF"  
                    link_cell.font = Font(color="0563C1", underline="single") 

                # 2. Format Priority Color
                priority_cell = worksheet.cell(row=row_num, column=priority_col_idx)
                priority_val = priority_cell.value
                if priority_val == "High":
                    priority_cell.fill = green_fill
                    priority_cell.font = green_text
                elif priority_val == "Medium":
                    priority_cell.fill = yellow_fill
                    priority_cell.font = yellow_text
                elif priority_val == "Low":
                    priority_cell.fill = red_fill
                    priority_cell.font = red_text

            # Loop 2: Apply Auto-Width, Text Wrapping, and Top Alignment
            for col_idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                
                for cell in col:
                    # Apply alignment to every cell
                    cell.alignment = wrap_top_align
                    
                    # Calculate string length for width
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set the column width. Cap it at 65 so long quotes don't stretch forever!
                adjusted_width = min(max_length + 2, 65)
                worksheet.column_dimensions[col_letter].width = adjusted_width

        # --- TAB 2: Discarded Log ---
        if not df_discarded.empty:
            df_discarded.to_excel(writer, sheet_name="Discarded Log", index=False)
            discard_ws = writer.sheets["Discarded Log"]
            
            # Apply auto-width and alignment to the discarded tab as well
            for col_idx, col in enumerate(discard_ws.columns, 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    cell.alignment = wrap_top_align
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                discard_ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    log.info(f"\n✅ Dashboard generated → {excel_path}")
# ─────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # ── Setup directories ────────────────────────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TRANSCRIPTS_DIR.mkdir(parents=True, exist_ok=True)

    # Initialize the NLP models before doing anything else
    nlp_engine_init()

    # ── Re-attach file log handler after dirs exist ──────────────────────────
    fh = logging.FileHandler(OUTPUT_DIR / "run.log")
    fh.setFormatter(logging.Formatter("%(asctime)s  [%(levelname)s]  %(message)s",
                                      "%H:%M:%S"))
    log.addHandler(fh)

    log.info("=" * 65)
    log.info("  CONCALL TRANSCRIPT ANALYZER  — Nifty 50 Test Run")
    log.info("=" * 65)

    # ── 1. Fetch universe ────────────────────────────────────────────────────
    # TEST MODE: Using Nifty 50 hardcoded list instead of the full NSE universe.
    # This skips the 2000-company yfinance loop entirely (~7 min saved).
    # To run the full universe, replace the line below with:
    #   df = fetch_nse_universe()
    #   df = enrich_market_cap(df)
    df = fetch_nse_universe()
    if df.empty:
        log.error("Could not fetch any companies. Exiting.")
        return

    df = enrich_market_cap(df)
    df = df.dropna(subset=["market_cap_crores"])
    log.info(f"Universe after dropping null market caps: {len(df)} companies\n")
    # ── 2. Run filter pipeline ───────────────────────────────────────────────
    df_filtered = run_filter_pipeline(df)
    if df_filtered.empty:
        log.error("Zero companies survived the filter pipeline. Exiting.")
        return

    # ── 3–5. Download transcripts (sequential — Playwright is single-threaded) ─
    fetcher = ScreenerTranscriptFetcher()
    
    results = []
    discarded = []  # NEW: Initialize the discarded list here!

    # Phase A: Use Playwright to find + download every PDF.
    # Must stay sequential — one browser, one page object.
    # Collect all successes into downloaded_jobs for Phase B.
    downloaded_jobs = []

    for idx, row in df_filtered.iterrows():
        symbol   = row["symbol"]
        name     = row["company_name"]
        mcap     = row["market_cap_crores"]

        log.info(f"[{idx+1}/{len(df_filtered)}]  {symbol}  ({name})")

        # -- Resolve BSE code from NSE symbol --------------------------------
        bse_code = fetcher.get_bse_code(symbol)
        if not bse_code:
            log.warning(f"  Could not resolve BSE code for {symbol} — skipping")
            discarded.append({"symbol": symbol, "company_name": name, "reason": "No BSE Code found"})
            continue

      # -- Find transcript URL ---------------------------------------------
        pdf_url = fetcher.fetch_latest_transcript_url(bse_code)
        if not pdf_url:
            log.info(f"  [{symbol}] Not found on Screener — trying NSE fallback …")
            pdf_url = fetcher.fetch_transcript_url_from_nse(symbol)

        if not pdf_url:
            log.info(f"  [{symbol}] No transcript found on Screener or NSE")
            discarded.append({"symbol": symbol, "company_name": name, "reason": "No Concall PDF found on Screener or NSE"})
            continue
        log.info(f"  PDF URL: {pdf_url}")

       # -- Download PDF (BSE via Screener) ---------------------------------
        pdf_path = TRANSCRIPTS_DIR / f"{symbol}_concall.pdf"
        if not pdf_path.exists():
            ok = fetcher.download_pdf(pdf_url, pdf_path)
            if not ok:
                # BSE/Screener blocked the download — try NSE fallback
                log.info(f"  [{symbol}] BSE download blocked — trying NSE fallback …")
                nse_url = fetcher.fetch_transcript_url_from_nse(symbol)
                if nse_url:
                    ok = fetcher.download_pdf(nse_url, pdf_path)
                    if ok:
                        pdf_url = nse_url
                        log.info(f"  [{symbol}] NSE fallback download succeeded")
                        _polite_sleep()
                    else:
                        discarded.append({"symbol": symbol, "company_name": name, "reason": "PDF Download Failed on BSE and NSE"})
                        continue
                else:
                    discarded.append({"symbol": symbol, "company_name": name, "reason": "PDF Download Failed on BSE, not found on NSE"})
                    continue
            else:
                _polite_sleep()
        else:
            log.info(f"  Using cached PDF for {symbol}")

        # -- Validate the downloaded PDF is real (not a blank/HTML page) -----
        # If the file is invalid, Layer 1 of extract_text_from_pdf will delete
        # it and return "". We detect that here and try NSE as a fallback.
        bse_pdf_valid = True
        try:
            with open(pdf_path, "rb") as f:
                raw = f.read(1024)
            if raw.lstrip()[:5] != b"%PDF-" or pdf_path.stat().st_size < 5_120:
                log.warning(f"  [{symbol}] BSE PDF is blank or invalid — trying NSE fallback …")
                pdf_path.unlink(missing_ok=True)
                bse_pdf_valid = False
        except Exception:
            bse_pdf_valid = False

        # -- NSE fallback: fetch a fresh PDF from NSE if BSE one was invalid --
        if not bse_pdf_valid:
            nse_url = fetcher.fetch_transcript_url_from_nse(symbol)
            if not nse_url:
                log.warning(f"  [{symbol}] No valid PDF found on BSE or NSE — skipping")
                discarded.append({"symbol": symbol, "company_name": name, "reason": "Blank PDF on BSE, not found on NSE"})
                continue
            ok = fetcher.download_pdf(nse_url, pdf_path)
            if not ok:
                discarded.append({"symbol": symbol, "company_name": name, "reason": "NSE PDF Download Failed"})
                continue
            pdf_url = nse_url
            _polite_sleep()
            log.info(f"  [{symbol}] NSE PDF downloaded successfully")

        # Queue for parse + score in Phase B
        downloaded_jobs.append({
            "symbol":            symbol,
            "company_name":      name,
            "market_cap_crores": mcap,
            "pdf_url":           pdf_url,
            "pdf_path":          pdf_path,
        })

    log.info(f"\n  Download phase complete. {len(downloaded_jobs)} PDFs queued for parsing.\n")

    # Phase B: Parse + score all PDFs in parallel.
    # extract_text_from_pdf and score_transcript have no shared mutable state —
    # safe to run concurrently. PDF_PARSE_WORKERS=4 is safe on any modern laptop.
    PDF_PARSE_WORKERS = 4

    def _parse_and_score(job: dict) -> dict:
        """Worker: parse one PDF and score it. Returns a result or discard dict."""
        sym      = job["symbol"]
        name_    = job["company_name"]
        mcap_    = job["market_cap_crores"]
        pdf_url_ = job["pdf_url"]
        pdf_path_= job["pdf_path"]

        text = extract_text_from_pdf(pdf_path_)
        if not text.strip():
            log.warning(f"  [{sym}] Empty text extracted from PDF")
            return {"_discard": True, "symbol": sym, "company_name": name_,
                    "reason": "PDF Text Extraction Failed"}

        score = score_transcript(text)
        log.info(f"  [{sym}] Score: {score['total_score']} categories "
                 f"| Priority: {score['priority']} "
                 f"| Matched: {', '.join(score['categories_found'])}")

        if score["priority"] == "Filtered Out":
            log.info(f"  [{sym}] → Below threshold — excluded from output")
            return {"_discard": True, "symbol": sym, "company_name": name_,
                    "reason": f"Only {score['total_score']} Categories Matched"}

        return {
            "_discard":          False,
            "symbol":            sym,
            "company_name":      name_,
            "market_cap_crores": mcap_,
            "transcript_url":    pdf_url_,
            **score,
        }

    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=PDF_PARSE_WORKERS) as executor:
        futures = [executor.submit(_parse_and_score, job) for job in downloaded_jobs]
        for future in concurrent.futures.as_completed(futures):
            try:
                outcome = future.result()
                if outcome.get("_discard"):
                    discarded.append({
                        "symbol":       outcome["symbol"],
                        "company_name": outcome["company_name"],
                        "reason":       outcome["reason"],
                    })
                else:
                    outcome.pop("_discard", None)
                    results.append(outcome)
            except Exception as e:
                log.warning(f"  Parse/score worker raised an error: {e}")

    # ── 6. Export ────────────────────────────────────────────────────────────
    OUTPUT_EXCEL = OUTPUT_DIR / f"concall_analysis_{datetime.today().strftime('%Y%m%d_%H%M%S')}"
    
    export_results_to_excel(results, discarded, OUTPUT_EXCEL)

    log.info(f"\nSummary (Nifty 50 Test Run — Revenue threshold: ≥{QUANTITATIVE_MIN_PCT:.0f}%):")
    log.info(f"  Companies in universe  : {len(df)}")
    log.info(f"  After filters          : {len(df_filtered)}")
    log.info(f"  With transcripts found : {len(downloaded_jobs)}")
    log.info(f"  In final output        : {len(results)}")
    high = sum(1 for r in results if r["priority"] == "High")
    med  = sum(1 for r in results if r["priority"] == "Medium")
    low  = sum(1 for r in results if r["priority"] == "Low")
    log.info(f"  High priority          : {high}")
    log.info(f"  Medium priority        : {med}")
    log.info(f"  Low priority           : {low}")
    log.info(f"  Discarded              : {len(discarded)}")
    log.info(f"  Output Excel           : {OUTPUT_EXCEL}.xlsx")
    log.info("=" * 65)


# ─────────────────────────────────────────────────────────────────────────────
# OFFLINE / DEMO MODE
# ─────────────────────────────────────────────────────────────────────────────
# When you want to test the pipeline without live HTTP calls, run:
#   python concall_analyzer.py --demo
#
# This feeds synthetic data through the scoring engine and exports a sample CSV.
# ─────────────────────────────────────────────────────────────────────────────

def demo_mode():
    """
    Test the pipeline end-to-end with synthetic transcript snippets.
    No internet access required.
    """
    log.info("=" * 65)
    log.info("  DEMO MODE — synthetic data, no HTTP calls")
    log.info("=" * 65)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    nlp_engine_init()

    synthetic = [
        {
            "symbol": "RELIANCE",
            "company_name": "Reliance Industries Ltd",
            "market_cap_crores": 1_800_000,
            "transcript": (
                "our capex guidance for the next fiscal is approximately 75000 crores "
                "focused on greenfield expansion in renewable energy and petrochemicals "
                "we aim for significant ebitda margin expansion we have strong revenue "
                "growth visibility backed by our order book in telecom and retail the "
                "management outlook remains positive we are targeting export markets in "
                "the middle east and europe as part of our global expansion strategy "
                "debt reduction continues to be a priority we plan on further "
                "deleveraging in fy25 r&d investments in new product categories are "
                "also ramping up"
            ),
        },
        {
            "symbol": "TATASTEEL",
            "company_name": "Tata Steel Ltd",
            "market_cap_crores": 220_000,
            "transcript": (
                "we are focused on cost reduction and operational efficiency across "
                "our plants capex spend this year is targeted at 12000 crores for "
                "capacity addition in pellet and downstream products guidance for "
                "ebitda is in the range of 10000 to 11000 crores we continue to "
                "address our debt situation with active repayment schedules"
            ),
        },
        {
            "symbol": "SMALLCAP",
            "company_name": "Tiny Corp Ltd",
            "market_cap_crores": 500,           # will be filtered by market cap
            "transcript": "we had a reasonable quarter sales were okay",
        },
    ]

    
    results = []
    discarded = []
    for co in synthetic:
        if co["market_cap_crores"] < MIN_MARKET_CAP_CRORES:
            log.info(f"  {co['symbol']} filtered out by market cap")
            continue
        text  = co["transcript"]
        score = score_transcript(text)
        log.info(f"  {co['symbol']}: score={score['total_score']}, "
                 f"priority={score['priority']}, "
                 f"categories={score['categories_found']}")
        if score["priority"] != "Filtered Out":
            co["transcript_url"] = "N/A (Demo)"
            results.append({**co, **score})

    export_results_to_excel(results, discarded, OUTPUT_DIR / "demo_output")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if "--demo" in sys.argv:
        demo_mode()
    else:
        main()