"""
=============================================================================
  CONCALL TRANSCRIPT ANALYZER FOR NSE/BSE LISTED COMPANIES
  ============================================================
  Author: Auto-generated modular pipeline script
  Purpose: Download, parse, and analyze earnings call transcripts
           for Indian-listed companies with expandable filtering.
=============================================================================

PIPELINE OVERVIEW:
  1. Fetch broad company universe  (NSE/BSE)
  2. Apply financial filters       (market cap, PE, sector, etc.)
  3. Download concall transcripts  (BSE XBRL / Screener.in)
  4. Parse PDF text
  5. Score against keyword dict
  6. Export prioritized CSV output

HOW TO ADD A NEW FILTER:
  - Write a function: filter_by_<your_criterion>(df) -> pd.DataFrame
  - Append it to FILTER_PIPELINE list near the bottom of the file.
  - That's it — the pipeline will pick it up automatically.
=============================================================================
"""
from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import numpy as np
import os
import re
import csv
import time
import logging
import random
import warnings
import requests
import io
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION  ← edit these variables freely
# ─────────────────────────────────────────────────────────────────────────────

# ── Market cap filter (in Crores INR) ────────────────────────────────────────
MIN_MARKET_CAP_CRORES = 5_000          # Only companies ≥ this pass through

# ── Keyword priority thresholds ──────────────────────────────────────────────
HIGH_PRIORITY_THRESHOLD   = 3.7          # ≥5 categories matched  → High
MEDIUM_PRIORITY_THRESHOLD = 2.8          # 4-5 categories matched → Medium
LOW_PRIORITY_THRESHOLD    = 2          # 2-3 categories matched → Low
                                       # <2 categories          → filtered out

# ── Output settings ──────────────────────────────────────────────────────────
OUTPUT_DIR    = Path("concall_output")
TRANSCRIPTS_DIR = OUTPUT_DIR / "transcripts"
OUTPUT_CSV    = OUTPUT_DIR / f"concall_analysis_{datetime.today().strftime('%Y%m%d')}.csv"

# ── Scraping politeness ──────────────────────────────────────────────────────
SLEEP_MIN_SECS = 1                     # Min sleep between HTTP requests
SLEEP_MAX_SECS = 3                     # Max sleep between HTTP requests

# ── Multi-concall / 3-year window settings ───────────────────────────────────
CONCALL_LOOKBACK_DAYS = 1_095          # 3 years (365 × 3)

# Year-bucket weights (must sum to 1.0):
#   Y1 = most recent 12 months, Y2 = months 13–24, Y3 = months 25–36
YEAR_WEIGHT_Y1 = 0.50
YEAR_WEIGHT_Y2 = 0.30
YEAR_WEIGHT_Y3 = 0.20

# ─────────────────────────────────────────────────────────────────────────────
# KEYWORD DICTIONARY
# ─────────────────────────────────────────────────────────────────────────────
# Each key is a CATEGORY (counts as 1 point max, regardless of how many
# synonyms are found).  Add/remove categories or synonyms freely.
# ─────────────────────────────────────────────────────────────────────────────

KEYWORDS_DICT = {
    "High Growth / Topline": [
        "topline growth", "revenue growth", "sales growth"
    ],
    "Order Visibility": [
        "order book", "new orders", "order inflow"
    ],
    "Margin Expansion": [
        "margin improvement", "ebit margin growth", "ebit margin expansion", "ebit margins growth", "ebit margins expansion",
        "ebitda margin growth", "ebitda margin expansion", "ebitda margins growth", "ebitda margins expansion",
        "better product mix", "premiumization"
    ],
    "Capacity Expansion": [
        "capacity expansion", "greenfield expansion", "brownfield expansion", "greenfield", 
        "brownfield", "new plant", "new facility", "capex for capacity", "capacity addition"
    ],
    "Integration & Efficiency": [
        "debottlenecking", "backward integration", "forward integration",
        "economies of scale", "improved operational efficiency"
    ],
    "Market Share": [
        "gaining market share", "improving market share", 
        "gaining wallet share", "improving wallet share"
    ],
    "Deleveraging": [
        "debt repayment", "deleveraging", "improved debt equity ratio", "debt replacement",
        "better rate of borrowing", "better rate of refinancing"
    ],
    "Cash Generation": [
        "net debt negative", "net cash positive", "fcf positive", "free cash flow positive"
    ],
    "Working Capital": [
        "working capital improvement", "working capital days reduced", 
        "cash conversion cycle improvement", "cash conversion cycle days reduced"
    ],
    "Return Ratios": [
        "roe improved", "roce improved", "roic improved", "roi improved",
        "return on equity improved", "return on capital employed improved", 
        "return on invested capital improved", "return on investment improved"
    ],
    "Corporate Action": [
        "acquisition", "merger", "demerger", "restructuring"
    ]
}

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING SETUP
# ─────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(OUTPUT_DIR / "run.log" if OUTPUT_DIR.exists()
                            else Path("concall_run.log")),
    ]
)
log = logging.getLogger("ConcallAnalyzer")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — FETCH COMPANY UNIVERSE
# ─────────────────────────────────────────────────────────────────────────────

def fetch_nse_universe() -> pd.DataFrame:
    """
    Download the master list of ALL listed equities on the NSE.
    """
    log.info("Fetching the complete NSE master company universe …")

    try:
        # The official NSE master list of all listed equities
        csv_url = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
        
        # We must spoof a browser user-agent, otherwise the NSE server blocks the download
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        
        s = requests.Session()
        s.headers.update(headers)
        r = s.get(csv_url, timeout=15)
        r.raise_for_status()
        
        # Read the downloaded CSV content
        df = pd.read_csv(io.StringIO(r.text))

        df.columns = df.columns.str.strip()
        
        # The CSV has columns like 'SYMBOL' and 'NAME OF COMPANY'
        df = df.rename(columns={
            "SYMBOL": "symbol",
            "NAME OF COMPANY": "company_name",
        })
        
        # Filter for standard equities (EQ) and book-entry (BE) to exclude bonds/ETFs
        df = df[df['SERIES'].isin(['EQ', 'BE'])].copy()
        
        # Initialize market cap to None so yfinance can fetch them all in the next step
        df["market_cap_crores"] = None          
        
        log.info(f"  ✓ Full NSE universe fetched: {len(df)} companies")
        return df[["symbol", "company_name", "market_cap_crores"]]

    except Exception as e:
        log.error(f"  Failed to fetch NSE master list: {e}")
        return pd.DataFrame(columns=["symbol", "company_name", "market_cap_crores"])


def enrich_market_cap(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fetch market caps using yfinance with parallel workers + one retry pass.
    """
    log.info("Enriching market cap data via yfinance (parallel) …")
    try:
        import yfinance as yf
    except ImportError:
        log.warning("  yfinance not installed. Run: pip install yfinance")
        df["market_cap_crores"] = 0
        return df

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def fetch_one(sym):
        try:
            info = yf.Ticker(f"{sym}.NS").fast_info
            mcap = getattr(info, "market_cap", None)
            return sym, (mcap / 1e7 if mcap else None)
        except Exception:
            return sym, None

    symbols = df["symbol"].tolist()
    results = {}

    # First pass — 20 workers
    with ThreadPoolExecutor(max_workers=20) as executor:
        for future in as_completed({executor.submit(fetch_one, s): s for s in symbols}):
            sym, mcap = future.result()
            results[sym] = mcap

    # Retry pass — only failed symbols, 5 workers (gentler)
    failed = [s for s, v in results.items() if v is None]
    if failed:
        log.info(f"  Retrying {len(failed)} failed symbols …")
        with ThreadPoolExecutor(max_workers=5) as executor:
            for future in as_completed({executor.submit(fetch_one, s): s for s in failed}):
                sym, mcap = future.result()
                if mcap is not None:
                    results[sym] = mcap

    df["market_cap_crores"] = df["symbol"].map(results).fillna(0)
    log.info(f"  ✓ Market cap enriched for {(df['market_cap_crores'] > 0).sum()} symbols")
    return df


def fetch_nifty50_universe() -> pd.DataFrame:
    """
    Return the Nifty 50 company list as a ready-to-use DataFrame.

    Used instead of fetch_nse_universe() + enrich_market_cap() when you want
    to test the scoring logic on just 50 companies without waiting for the
    full 2000-company yfinance enrichment loop.

    Market caps are approximate (order-of-magnitude) — sufficient for the
    market cap filter to pass all Nifty 50 companies through (they all exceed
    5,000 Cr comfortably). Screener uses NSE symbols directly, so no BSE
    code resolution is needed.

    To switch back to the full NSE run, replace fetch_nifty50_universe() in
    main() with fetch_nse_universe() + enrich_market_cap().
    """
    log.info("TEST MODE: Using hardcoded Nifty 50 universe (skipping yfinance loop) …")

    # Current Nifty 50 constituents (NSE symbols).
    # Note: HDFC Ltd merged into HDFCBANK in 2023 and is no longer listed.
    # JIOFIN, ETERNAL (Zomato), BEL, TRENT are recent additions.
    nifty50_data = [
        ("ADANIENT",   "Adani Enterprises Ltd",                      250_000),
        ("ADANIPORTS",  "Adani Ports & SEZ Ltd",                      310_000),
        ("APOLLOHOSP",  "Apollo Hospitals Enterprise Ltd",             100_000),
        ("ASIANPAINT",  "Asian Paints Ltd",                            195_000),
        ("AXISBANK",    "Axis Bank Ltd",                               370_000),
        ("BAJAJ-AUTO",  "Bajaj Auto Ltd",                              225_000),
        ("BAJAJFINSV",  "Bajaj Finserv Ltd",                           270_000),
        ("BAJFINANCE",  "Bajaj Finance Ltd",                           440_000),
        ("BEL",         "Bharat Electronics Ltd",                      200_000),
        ("BHARTIARTL",  "Bharti Airtel Ltd",                           950_000),
        ("BRITANNIA",   "Britannia Industries Ltd",                    115_000),
        ("CIPLA",       "Cipla Ltd",                                   120_000),
        ("COALINDIA",   "Coal India Ltd",                              240_000),
        ("DIVISLAB",    "Divi's Laboratories Ltd",                     130_000),
        ("DRREDDY",     "Dr. Reddy's Laboratories Ltd",                140_000),
        ("EICHERMOT",   "Eicher Motors Ltd",                           130_000),
        ("ETERNAL",     "Eternal Ltd (Zomato)",                        220_000),
        ("GRASIM",      "Grasim Industries Ltd",                       175_000),
        ("HCLTECH",     "HCL Technologies Ltd",                        440_000),
        ("HDFCBANK",    "HDFC Bank Ltd",                             1_450_000),
        ("HDFCLIFE",    "HDFC Life Insurance Company Ltd",             140_000),
        ("HEROMOTOCO",  "Hero MotoCorp Ltd",                           100_000),
        ("HINDALCO",    "Hindalco Industries Ltd",                     200_000),
        ("HINDUNILVR",  "Hindustan Unilever Ltd",                      530_000),
        ("ICICIBANK",   "ICICI Bank Ltd",                              900_000),
        ("INFY",        "Infosys Ltd",                                 680_000),
        ("ITC",         "ITC Ltd",                                     500_000),
        ("JIOFIN",      "Jio Financial Services Ltd",                  195_000),
        ("JSWSTEEL",    "JSW Steel Ltd",                               240_000),
        ("KOTAKBANK",   "Kotak Mahindra Bank Ltd",                     390_000),
        ("LT",          "Larsen & Toubro Ltd",                         500_000),
        ("M&M",         "Mahindra & Mahindra Ltd",                     370_000),
        ("MARUTI",      "Maruti Suzuki India Ltd",                     360_000),
        ("NESTLEIND",   "Nestle India Ltd",                            225_000),
        ("NTPC",        "NTPC Ltd",                                    370_000),
        ("ONGC",        "Oil & Natural Gas Corporation Ltd",           310_000),
        ("POWERGRID",   "Power Grid Corporation of India Ltd",         290_000),
        ("RELIANCE",    "Reliance Industries Ltd",                   1_800_000),
        ("SBILIFE",     "SBI Life Insurance Company Ltd",              145_000),
        ("SBIN",        "State Bank of India",                         700_000),
        ("SUNPHARMA",   "Sun Pharmaceutical Industries Ltd",           390_000),
        ("TATACONSUM",  "Tata Consumer Products Ltd",                  115_000),
        ("TATAMOTORS",  "Tata Motors Ltd",                             310_000),
        ("TATASTEEL",   "Tata Steel Ltd",                              180_000),
        ("TCS",         "Tata Consultancy Services Ltd",             1_400_000),
        ("TECHM",       "Tech Mahindra Ltd",                           165_000),
        ("TITAN",       "Titan Company Ltd",                           275_000),
        ("TRENT",       "Trent Ltd",                                   175_000),
        ("ULTRACEMCO",  "UltraTech Cement Ltd",                        300_000),
        ("WIPRO",       "Wipro Ltd",                                   265_000),
    ]

    df = pd.DataFrame(nifty50_data, columns=["symbol", "company_name", "market_cap_crores"])
    log.info(f"  ✓ Nifty 50 universe loaded: {len(df)} companies")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — MODULAR FILTERING PIPELINE
# ─────────────────────────────────────────────────────────────────────────────
# Each filter function MUST:
#   • Accept a pd.DataFrame  (columns: symbol, company_name, market_cap_crores)
#   • Return a filtered pd.DataFrame
#   • Log how many rows were retained
#
# To add a new filter: write the function, append it to FILTER_PIPELINE list.
# ─────────────────────────────────────────────────────────────────────────────

def filter_by_market_cap(df: pd.DataFrame) -> pd.DataFrame:
    """[ACTIVE] Keep companies with market cap ≥ MIN_MARKET_CAP_CRORES."""
    before = len(df)
    df = df[df["market_cap_crores"] >= MIN_MARKET_CAP_CRORES].copy()
    log.info(f"  filter_by_market_cap  → {len(df)}/{before} companies retained "
             f"(≥ ₹{MIN_MARKET_CAP_CRORES:,} Cr)")
    return df


# ── TEMPLATE FILTERS — uncomment & implement when needed ─────────────────────

# def filter_by_pe_ratio(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep companies with PE ratio ≤ MAX_PE."""
#     MAX_PE = 40
#     # TODO: add 'pe_ratio' column to df before this filter runs
#     before = len(df)
#     df = df[df["pe_ratio"] <= MAX_PE].copy()
#     log.info(f"  filter_by_pe_ratio    → {len(df)}/{before} retained (PE ≤ {MAX_PE})")
#     return df


# def filter_by_sector(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep only companies in ALLOWED_SECTORS."""
#     ALLOWED_SECTORS = {"FMCG", "Pharmaceuticals", "Technology", "Banking"}
#     # TODO: add 'sector' column to df before this filter runs
#     before = len(df)
#     df = df[df["sector"].isin(ALLOWED_SECTORS)].copy()
#     log.info(f"  filter_by_sector      → {len(df)}/{before} retained")
#     return df


# def filter_by_revenue_growth(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep companies with YoY revenue growth ≥ threshold."""
#     MIN_REVENUE_GROWTH_PCT = 10
#     before = len(df)
#     df = df[df["revenue_growth_pct"] >= MIN_REVENUE_GROWTH_PCT].copy()
#     log.info(f"  filter_by_revenue_growth → {len(df)}/{before} retained")
#     return df


# def filter_by_debt_to_equity(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Remove highly leveraged companies."""
#     MAX_DE_RATIO = 1.5
#     before = len(df)
#     df = df[df["de_ratio"] <= MAX_DE_RATIO].copy()
#     log.info(f"  filter_by_de_ratio    → {len(df)}/{before} retained")
#     return df


# def filter_by_promoter_holding(df: pd.DataFrame) -> pd.DataFrame:
#     """[TEMPLATE] Keep companies where promoter holding ≥ threshold."""
#     MIN_PROMOTER_HOLDING_PCT = 30
#     before = len(df)
#     df = df[df["promoter_holding_pct"] >= MIN_PROMOTER_HOLDING_PCT].copy()
#     log.info(f"  filter_by_promoter    → {len(df)}/{before} retained")
#     return df


# ── Register active filters here (order matters) ─────────────────────────────
FILTER_PIPELINE = [
    filter_by_market_cap,
    # filter_by_pe_ratio,
    # filter_by_sector,
    # filter_by_revenue_growth,
    # filter_by_debt_to_equity,
    # filter_by_promoter_holding,
]


def run_filter_pipeline(df: pd.DataFrame) -> pd.DataFrame:
    """Run df through every filter in FILTER_PIPELINE sequentially."""
    log.info(f"Starting filter pipeline with {len(df)} companies …")
    for fn in FILTER_PIPELINE:
        df = fn(df)
        if df.empty:
            log.warning("  Pipeline produced 0 results — check filter thresholds!")
            break
    log.info(f"  → {len(df)} companies passed all filters\n")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — TRANSCRIPT DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────
# DESIGN NOTE ON ANTI-SCRAPING:
#
#   BSE India & Screener.in use Cloudflare + dynamic tokens.  Two safe paths:
#
#   OPTION A (recommended): BSE XBRL open API
#     BSE provides an official XML/JSON filings feed. Concall transcripts
#     are filed under 'Corporate Announcements' (categoryId=6 or subcategory
#     "Conference Call" / "Investor Presentation"). No auth needed; just
#     rotate User-Agents and respect rate limits.
#
#   OPTION B: Screener.in unofficial wrapper
#     Library: https://github.com/pratik2315/screener-python (community)
#     Install: pip install screener-python
#     Caveat: token-based, may break on UI changes — use as fallback only.
#
#   OPTION C: NSE XBRL feed (alternative to BSE)
#     NSE also publishes announcements via:
#     https://www.nseindia.com/api/corporate-announcements
# ─────────────────────────────────────────────────────────────────────────────

# Realistic browser User-Agent pool for rotation
_USER_AGENTS = [
    ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
     "AppleWebKit/537.36 (KHTML, like Gecko) "
     "Chrome/124.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) "
     "AppleWebKit/605.1.15 (KHTML, like Gecko) "
     "Version/17.4 Safari/605.1.15"),
    ("Mozilla/5.0 (X11; Linux x86_64; rv:125.0) "
     "Gecko/20100101 Firefox/125.0"),
]


def _get_headers(referer: str = "https://www.bseindia.com") -> dict:
    """Return headers that mimic a real browser, with rotated User-Agent."""
    return {
        "User-Agent":      random.choice(_USER_AGENTS),
        "Accept":          "application/json, text/plain, */*",
        "Accept-Language": "en-IN,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer":         referer,
        "DNT":             "1",
        "Connection":      "keep-alive",
    }


def _polite_sleep():
    """Sleep a random interval to avoid getting blocked."""
    secs = random.uniform(SLEEP_MIN_SECS, SLEEP_MAX_SECS)
    time.sleep(secs)


import json
import time
import logging
import requests
import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

log = logging.getLogger("ConcallAnalyzer")

class ScreenerTranscriptFetcher:
    def __init__(self):
        log.info("  Booting up stealth browser engine for Screener...")
        self.playwright = sync_playwright().start()
        
        # NOTE: If you still get blocked, change headless=True to headless=False 
        # so you can see the browser and manually click any Cloudflare boxes!
        self.browser = self.playwright.chromium.launch(headless=False) 
        self.context = self.browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        self.page = self.context.new_page()
        self.nse_page = self.context.new_page()

        # Warm up the browser on Screener's homepage to grab initial cookies
        try:
            self.page.goto("https://www.screener.in", timeout=60000)
            self.page.wait_for_timeout(3000)
        except Exception as e:
            log.debug(f"  Browser warmup timeout: {e}")

    def __del__(self):
        # Clean up the browser when the script finishes
        try:
            if hasattr(self, 'browser'):
                self.browser.close()
            if hasattr(self, 'playwright'):
                self.playwright.stop()
        except Exception:
            pass

    def get_bse_code(self, nse_symbol: str):
        # We don't need BSE codes for Screener! Just return the NSE symbol.
        return nse_symbol.strip().upper()

    def fetch_all_transcript_urls(self, symbol: str,
                                   lookback_days: int = CONCALL_LOOKBACK_DAYS
                                   ) -> list[tuple[datetime, str]]:
        """
        Scrape ALL transcript/concall PDF links from Screener within the
        lookback window.

        Returns a list of (date, url) tuples sorted newest-first.
        Falls back to an empty list on any error.

        Screener renders each concall link inside an <li> element like:
          <li>
            <span class="date">Nov 2024</span>
            <a href="/api/media/concall/...pdf">Transcript</a>
          </li>
        We walk every such anchor and parse its sibling date span.
        If date parsing fails we still include the link (date = today).
        """
        from datetime import datetime, timedelta
        cutoff = datetime.today() - timedelta(days=lookback_days)
        results: list[tuple[datetime, str]] = []

        url = f"https://www.screener.in/company/{symbol}/consolidated/"
        try:
            self.page.goto(url, timeout=30000)
            self.page.wait_for_timeout(2000)

            page_title = self.page.title()
            if "Just a moment" in page_title or "Cloudflare" in page_title:
                log.warning(f"  [!] Cloudflare triggered for {symbol} — please click the checkbox …")
                self.page.wait_for_timeout(12000)

            html = self.page.content()
            soup = BeautifulSoup(html, "html.parser")

            # ── Strategy 1: walk every <a> that contains "Transcript" or "Concall" ──
            anchors = soup.find_all(
                "a",
                string=lambda t: t and ("Transcript" in t or "Concall" in t or "transcript" in t),
            )

            for a in anchors:
                href = a.get("href", "")
                if not href:
                    continue
                if href.startswith("/"):
                    href = "https://www.screener.in" + href

                # Try to find a nearby date: look in parent <li> or <div> for a
                # <span class="date"> or any text that looks like "Nov 2024" / "Q3 FY25"
                link_date = datetime.today()           # default if we can't parse
                parent = a.find_parent(["li", "div", "tr"])
                if parent:
                    date_span = parent.find("span", class_=lambda c: c and "date" in c.lower())
                    date_text = date_span.get_text(strip=True) if date_span else parent.get_text(" ", strip=True)
                    link_date = _parse_concall_date(date_text, fallback=datetime.today())

                if link_date >= cutoff:
                    results.append((link_date, href))

            # ── Strategy 2: look inside a "Documents" / "Concalls" section ──────
            # Some Screener pages list PDFs in a structured documents block
            if not results:
                doc_section = soup.find(
                    lambda tag: tag.name in ("section", "div") and
                    any(kw in (tag.get("id", "") + " " + " ".join(tag.get("class", []))).lower()
                        for kw in ("document", "concall", "transcript", "annual"))
                )
                if doc_section:
                    for a in doc_section.find_all("a", href=True):
                        href = a["href"]
                        if not any(kw in href.lower() for kw in ("concall", "transcript", ".pdf")):
                            continue
                        if href.startswith("/"):
                            href = "https://www.screener.in" + href
                        results.append((datetime.today(), href))

        except Exception as e:
            log.debug(f"  Screener multi-fetch error for {symbol}: {e}")

        # Deduplicate URLs, sort newest-first
        seen: set[str] = set()
        unique: list[tuple[datetime, str]] = []
        for dt, u in sorted(results, key=lambda x: x[0], reverse=True):
            if u not in seen:
                seen.add(u)
                unique.append((dt, u))

        log.info(f"  [{symbol}] Screener returned {len(unique)} transcript links within 3-year window")
        return unique

    def fetch_all_from_nse(self, symbol: str,
                            lookback_days: int = CONCALL_LOOKBACK_DAYS
                            ) -> list[tuple[datetime, str]]:
        """
        NSE fallback: return ALL concall transcript PDFs within the lookback
        window from the NSE corporate-announcements API.

        Returns list of (date, url) tuples sorted newest-first.
        """
        from datetime import datetime, timedelta
        results: list[tuple[datetime, str]] = []

        try:
            self.nse_page.goto("https://www.nseindia.com", timeout=30000)
            self.nse_page.wait_for_timeout(2000)

            api_url = (
                f"https://www.nseindia.com/api/corporate-announcements"
                f"?index=equities&symbol={symbol}"
            )

            response = self.nse_page.evaluate(f"""
                async () => {{
                    const r = await fetch("{api_url}", {{
                        headers: {{
                            "Accept": "application/json",
                            "Referer": "https://www.nseindia.com/"
                        }}
                    }});
                    return await r.json();
                }}
            """)

            if not response or not isinstance(response, list):
                log.warning(f"  [{symbol}] NSE returned no data")
                return []

            cutoff = datetime.today() - timedelta(days=lookback_days)

            for item in response:
                desc         = (item.get("desc", "") or "").lower()
                attchmnt_text= (item.get("attchmntText", "") or "").lower()
                an_dt_str    = item.get("an_dt", "")

                if not any(kw in desc for kw in (
                    "concall", "transcript", "conference call",
                    "earnings call", "con. call"
                )):
                    continue

                if any(kw in attchmnt_text for kw in (
                    "schedule of meet", "schedule of analyst",
                    "intimation of meet", "inform the exchange about schedule"
                )):
                    continue

                try:
                    an_dt = datetime.strptime(an_dt_str.split(" ")[0], "%d-%b-%Y")
                except Exception:
                    continue

                if an_dt < cutoff:
                    continue

                pdf_file = item.get("attchmntFile", "")
                if pdf_file:
                    pdf_url = (
                        pdf_file if pdf_file.startswith("http")
                        else f"https://nsearchives.nseindia.com/corporate/{pdf_file}"
                    )
                    results.append((an_dt, pdf_url))
                    log.info(f"  [{symbol}] NSE link: {pdf_file} ({an_dt.strftime('%b %Y')})")

        except Exception as e:
            log.debug(f"  [{symbol}] NSE multi-fetch error: {e}")

        # Sort newest-first, deduplicate
        seen: set[str] = set()
        unique: list[tuple[datetime, str]] = []
        for dt, u in sorted(results, key=lambda x: x[0], reverse=True):
            if u not in seen:
                seen.add(u)
                unique.append((dt, u))

        log.info(f"  [{symbol}] NSE returned {len(unique)} transcript links within 3-year window")
        return unique

    # ── Keep the old single-URL method as a thin wrapper for backward compat ──
    def fetch_transcript_url_from_nse(self, symbol: str, lookback_days: int = 180):
        results = self.fetch_all_from_nse(symbol, lookback_days)
        return results[0][1] if results else None

    def download_pdf(self, pdf_url: str, save_path: Path):
        """
        Download the PDF using stealth headers to bypass BSE/Screener blocks.
        """
        # THE FIX: Tell the server we are a real Chrome browser coming from BSE
        stealth_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/pdf,application/xhtml+xml,text/html;q=0.9,*/*;q=0.8",
            "Referer": "https://www.bseindia.com/"  # Critical for BSE links!
        }
        
        try:
            # Pass the headers into the request
            response = requests.get(pdf_url, headers=stealth_headers, timeout=15)
            response.raise_for_status()
            
            # Check if we accidentally got an HTML page anyway
            if b"<!DOC" in response.content[:10].upper() or b"<HTML" in response.content[:10].upper():
                log.warning("Server blocked the PDF and sent an HTML page instead.")
                return False

            # Save the real PDF
            with open(save_path, "wb") as f:
                f.write(response.content)
            return True
            
        except Exception as e:
            log.error(f"Failed to download PDF: {e}")
            return False
# ─────────────────────────────────────────────────────────────────────────────
# MULTI-CONCALL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_concall_date(text: str, fallback: "datetime") -> "datetime":
    """
    Try to extract a date from text like "Nov 2024", "Q3 FY25", "15 Jan 2025",
    "October 2024", etc.  Returns fallback if nothing parses.
    """
    import re as _re
    from datetime import datetime as _dt

    text = text.strip()

    # Pattern: "15 Jan 2025" or "Jan 15, 2025"
    for fmt in ("%d %b %Y", "%b %d, %Y", "%d %B %Y", "%B %d, %Y"):
        try:
            return _dt.strptime(text[:len(fmt) + 2].strip(), fmt)
        except ValueError:
            pass

    # Pattern: "Nov 2024" / "November 2024"
    m = _re.search(r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|"
                   r"Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|"
                   r"Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+(\d{4})", text, _re.I)
    if m:
        try:
            return _dt.strptime(f"01 {m.group(1)[:3]} {m.group(2)}", "%d %b %Y")
        except ValueError:
            pass

    # Pattern: "Q3 FY25" / "Q1FY2025"
    m2 = _re.search(r"Q([1-4])\s*FY\s*(\d{2,4})", text, _re.I)
    if m2:
        quarter = int(m2.group(1))
        yr_raw  = m2.group(2)
        year    = int(yr_raw) if len(yr_raw) == 4 else 2000 + int(yr_raw)
        # Indian FY: Q1=Apr, Q2=Jul, Q3=Oct, Q4=Jan
        month_map = {1: 4, 2: 7, 3: 10, 4: 1}
        month = month_map[quarter]
        adj_year = year if quarter != 4 else year - 1
        try:
            return _dt(adj_year, month, 1)
        except ValueError:
            pass

    return fallback


def compute_weighted_score(concall_records: list[dict]) -> float:
    """
    Compute a weighted score (0–11) across multiple concall records.

    Parameters
    ----------
    concall_records : list of dicts, each with keys:
        "date"        (datetime)
        "total_score" (int, 0–11)

    Weighting logic
    ---------------
    • Year buckets (from today backwards):
        Y1 = 0–365 days  → 50% of total weight
        Y2 = 366–730 days → 30%
        Y3 = 731–1095 days → 20%
    • Within each bucket, concalls are ranked newest-first and assigned
      a linear share proportional to their rank:
        e.g. 4 concalls in Y1 → weights 4/10, 3/10, 2/10, 1/10
        (sum = 1.0 within the bucket before applying bucket weight)
    • Final score = sum of (normalised_weight × total_score) across all records,
      capped at 11.0

    If a bucket has zero records its weight is redistributed proportionally
    to the remaining non-empty buckets.
    """
    from datetime import datetime, timedelta

    if not concall_records:
        return 0.0

    today = datetime.today()
    b1_end = today - timedelta(days=0)
    b1_cut = today - timedelta(days=365)
    b2_cut = today - timedelta(days=730)
    b3_cut = today - timedelta(days=1095)

    # Assign each record to a bucket (1, 2, or 3)
    buckets: dict[int, list[dict]] = {1: [], 2: [], 3: []}
    for rec in concall_records:
        age = (today - rec["date"]).days
        if age <= 365:
            buckets[1].append(rec)
        elif age <= 730:
            buckets[2].append(rec)
        else:
            buckets[3].append(rec)

    # Sort each bucket newest-first
    for b in buckets.values():
        b.sort(key=lambda r: r["date"], reverse=True)

    # Base bucket weights
    raw_weights = {1: YEAR_WEIGHT_Y1, 2: YEAR_WEIGHT_Y2, 3: YEAR_WEIGHT_Y3}

    # Redistribute weight of empty buckets to non-empty ones proportionally
    empty   = [b for b, recs in buckets.items() if not recs]
    nonempty= [b for b, recs in buckets.items() if recs]
    if empty and nonempty:
        freed = sum(raw_weights[b] for b in empty)
        total_ne = sum(raw_weights[b] for b in nonempty)
        for b in nonempty:
            raw_weights[b] += freed * (raw_weights[b] / total_ne)

    weighted_sum = 0.0
    for bucket_id, recs in buckets.items():
        if not recs:
            continue
        bucket_w = raw_weights[bucket_id]
        n = len(recs)
        # Intra-bucket rank weights: newest = n, oldest = 1, normalised
        rank_total = n * (n + 1) / 2
        for rank_idx, rec in enumerate(recs):       # recs already sorted newest-first
            rank_val = n - rank_idx                 # newest gets n, oldest gets 1
            intra_w  = rank_val / rank_total        # normalised within bucket
            combined_w = bucket_w * intra_w
            weighted_sum += combined_w * rec["total_score"]

    return round(min(weighted_sum, 11.0), 2)


def _fy_label(dt: "datetime") -> str:
    """
    Return a human-readable Indian FY quarter label for a given date.
    e.g.  datetime(2024, 11, 1)  →  "Q3 FY25"
          datetime(2025, 1, 1)   →  "Q4 FY25"
    """
    from datetime import datetime
    m, y = dt.month, dt.year
    # Indian FY starts April 1 — FY label is the year it ends
    fy = y + 1 if m >= 4 else y
    # Quarter mapping: Apr-Jun=Q1, Jul-Sep=Q2, Oct-Dec=Q3, Jan-Mar=Q4
    if   m in (4, 5, 6):   q = "Q1"
    elif m in (7, 8, 9):   q = "Q2"
    elif m in (10, 11, 12): q = "Q3"
    else:                   q = "Q4"
    return f"{q} FY{str(fy)[-2:]}"


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — PDF TEXT PARSING
# ─────────────────────────────────────────────────────────────────────────────

def extract_text_from_pdf(pdf_path: Path) -> str:
    """
    Extract raw text from a PDF with validation and dual-parser fallback.
    """
    # ── Layer 1: Validate ────────────────────────────────────────────────────
    try:
        file_size = pdf_path.stat().st_size
        if file_size < 5_120:
            log.warning(f"  [{pdf_path.stem}] Too small ({file_size} bytes) — deleting cache.")
            pdf_path.unlink(missing_ok=True)
            return ""
        with open(pdf_path, "rb") as f:
            raw = f.read(1024)
        header_check = raw.lstrip()[:5]
        if header_check != b"%PDF-":
            log.warning(f"  [{pdf_path.stem}] Not a valid PDF (got {header_check}) — deleting cache.")
            pdf_path.unlink(missing_ok=True)
            return ""
    except Exception as e:
        log.warning(f"  [{pdf_path.stem}] Validation error: {e}")
        return ""

    # ── Layer 2: pdfplumber ──────────────────────────────────────────────────
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            pages_text = []
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    pages_text.append(t)
            text = "\n".join(pages_text)
        if text.strip():
            return text
    except ImportError:
        pass
    except Exception as e:
        log.debug(f"  [{pdf_path.stem}] pdfplumber error: {e}")

    # ── Layer 3: pypdf ───────────────────────────────────────────────────────
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        pages_text = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                pages_text.append(t)
        text = "\n".join(pages_text)
        if text.strip():
            return text
    except ImportError:
        log.error("  pypdf not installed. Run: pip install pypdf")
    except Exception as e:
        log.debug(f"  [{pdf_path.stem}] pypdf error: {e}")

    return ""

def clean_text(text: str) -> str:
    """
    Normalise transcript text for keyword scanning:
      - Lowercase
      - Collapse whitespace / newlines
      - Remove non-alphanumeric characters except spaces and hyphens
    """
    text = text.lower()
    text = re.sub(r"[\r\n\t]+", " ", text)          # flatten newlines
    text = re.sub(r"[^a-z0-9 \-]", " ", text)       # strip punctuation
    text = re.sub(r"\s{2,}", " ", text).strip()      # collapse spaces
    return text


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — KEYWORD SCORING
# ─────────────────────────────────────────────────────────────────────────────

# Layer 1 — minimum % threshold for the Revenue Growth Rule
# Set to 10.0 for the Nifty 50 test run (large-caps grow slower; 20% would miss most).
# Change back to 20.0 for the full NSE universe run.
QUANTITATIVE_MIN_PCT: float = 20.0

# Layer 2 — minimum cosine similarity score to count as a semantic match
SEMANTIC_SIMILARITY_THRESHOLD: float = 0.75

# Layer 3 — DistilBERT negative-score threshold (0–1 scale).
# Sentences with a NEGATIVE score above this are discarded.
SENTIMENT_NEGATIVE_THRESHOLD: float = 0.60

# Maximum number of quote snippets written to context_snippets (Excel cell limit)
MAX_SNIPPETS: int = 7

# Sentence minimum length filter (characters)
MIN_SENTENCE_LEN: int = 25


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 1 ─ QUANTITATIVE EXTRACTION PATTERNS
# ─────────────────────────────────────────────────────────────────────────────
# Each category maps to a list of regex patterns.  Every pattern MUST contain
# exactly one capture group that extracts the percentage number.
# Patterns are tried in order; the first match wins.
# ─────────────────────────────────────────────────────────────────────────────

QUANTITATIVE_PATTERNS: dict[str, list[str]] = {

    "High Growth / Topline": [
        # "revenue/sales grew/increased/jumped by 25%"
        r"(?:topline|top[\-\s]line|revenue|net\s+sales?|sales?|turnover)"
        r"[\w\s,;]+?(?:grew?|grow(?:ing|th)|increas(?:ed?|ing)|jump(?:ed)?|"
        r"surged?|rose?|up|higher)\s+(?:by\s+)?(\d+(?:\.\d+)?)\s*%",
        # "revenue growth of 25%" / "sales growth of 8%" — noun phrase form
        # MUST be here so Layer 1 intercepts it and enforces the 20% threshold
        # before it falls through to the keyword fallback which has no % check.
        r"(?:topline|top[\-\s]line|revenue|net\s+sales?|sales?|turnover)"
        r"\s+growth\s+of\s+(\d+(?:\.\d+)?)\s*%",
        # "25% revenue growth / 25% growth in revenue"
        r"(\d+(?:\.\d+)?)\s*%\s+(?:revenue|sales?|topline|top[\-\s]line|"
        r"turnover)\s+(?:growth|increase|jump|surge)",
        r"(\d+(?:\.\d+)?)\s*%\s+(?:growth|increase|jump|surge)\s+in\s+"
        r"(?:revenue|sales?|topline)",
        # "achieved/delivered/posted 30% revenue growth"
        r"(?:achiev|record|report|deliver|post|clock)(?:ed?|ing)\s+"
        r"(\d+(?:\.\d+)?)\s*%\s+(?:revenue|sales?|topline|top[\-\s]line)",
        # "22% yoy revenue growth"
        r"(\d+(?:\.\d+)?)\s*%\s+(?:yoy|y[\-\s]o[\-\s]y|year[\-\s]on[\-\s]year|"
        r"qoq|q[\-\s]o[\-\s]q|quarter[\-\s]on[\-\s]quarter)\s+"
        r"(?:revenue|sales?|growth|topline)",
    ],

    "Order Visibility": [
        # "order inflows of INR X crores" — no % threshold applied to this category
        # because order book sizes are typically in absolute INR, not %.
        # We keep it in QUANTITATIVE_PATTERNS as an optional trigger.
        # Use a dummy pattern that intentionally never fires (order book
        # matching is handled fully by the keyword layer below).
        r"order\s+(?:book|inflow|backlog)\s+(?:at|of|stands?\s+at)\s+"
        r"(?:inr|rs\.?|₹)?\s*(\d+(?:\.\d+)?)\s*(?:crore|cr|lakh|mn|bn)",
    ],

    "Margin Expansion": [
        # "EBITDA margins expanded/improved by 200 bps to 22%"
        r"(?:ebitda|ebit|gross|operating|net|pbt)\s+margins?\s+"
        r"[\w\s,;]+?(?:expand|improv|increas|up|higher|widen|grew?)\s+"
        r"(?:by\s+)?(\d+(?:\.\d+)?)\s*%",
        # "margins expanded 150 bps" (basis points → treated as pct for threshold)
        r"margins?\s+[\w\s]+?(?:expand|improv|up|widen)\s+(?:by\s+)?(\d+)\s*bps",
        # "margin now at 25%" — absolute margin level used as proxy
        r"(?:ebitda|ebit|gross|operating)\s+margins?\s+(?:of|at|to)\s+"
        r"(\d+(?:\.\d+)?)\s*%",
    ],

    "Return Ratios": [
        # "ROCE improved to 22%"
        r"(?:roce|roe|roic|roi|return\s+on\s+(?:capital|equity|invest))"
        r"[\w\s,;]+?(?:improv|increas|expand|up|higher|rose?|grew?)\s+"
        r"(?:by\s+)?(?:to\s+)?(\d+(?:\.\d+)?)\s*%",
        r"(?:roce|roe|roic|roi)\s+(?:of|at|now|stands?\s+at)\s+"
        r"(\d+(?:\.\d+)?)\s*%",
    ],

    "Market Share": [
        # "market share grew from 12% to 17%"
        r"market\s+share\s+[\w\s,;]+?(?:grew?|up|higher|from\s+\d+%?\s+to)\s+"
        r"(?:to\s+)?(\d+(?:\.\d+)?)\s*%",
        r"(?:market|wallet)\s+share\s+(?:of|at|to)\s+(\d+(?:\.\d+)?)\s*%",
    ],
}


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 2 ─ SEMANTIC SEARCH CONCEPT LIBRARY
# ─────────────────────────────────────────────────────────────────────────────
# These are the "concept anchors" that will be encoded into embedding vectors.
# Add richer paraphrase variations to improve recall.
# ─────────────────────────────────────────────────────────────────────────────

SEMANTIC_CONCEPTS: dict[str, list[str]] = {

    "Margin Expansion": [
        "EBITDA margin expansion and profitability improvement",
        "gross margin improved due to better product mix",
        "operating leverage driving margin expansion",
        "margins expanded as fixed costs got absorbed on higher revenues",
        "cost efficiencies leading to margin improvement",
        "better realisations improving profitability",
        "premium product mix driving margin accretion",
    ],

    "Capacity Expansion": [
        "new greenfield plant commissioned and ramping up",
        "brownfield capacity expansion adding production capacity",
        "capex investment in new manufacturing facility",
        "capacity addition to meet growing demand",
        "new plant going on stream increasing installed capacity",
        "expanding production footprint to scale operations",
    ],

    "Deleveraging": [
        "significant debt repayment reducing leverage on balance sheet",
        "net debt to EBITDA ratio coming down",
        "balance sheet deleveraging with debt reduction",
        "company becoming debt free after debt repayment",
        "leverage ratio improved after repaying term loans",
        "strong free cash flows used for debt reduction",
    ],

    "Integration & Efficiency": [
        "backward integration reducing raw material dependency",
        "forward integration into higher value-added products",
        "debottlenecking increasing throughput without capex",
        "economies of scale improving operational efficiency",
        "process automation driving cost savings",
        "vertical integration improving margin and supply security",
    ],

    "Cash Generation": [
        "strong free cash flow generation supporting growth investments",
        "company turned net cash positive with no net debt",
        "FCF positive and self-funding growth capex",
        "robust operating cash flows funding capex and debt repayment",
        "net debt negative balance sheet with cash surplus",
    ],

    "Working Capital": [
        "working capital days reduced improving cash conversion",
        "inventory days and receivable days improved",
        "cash conversion cycle shortened due to better collections",
        "debtors days reduced as collection efficiency improved",
        "working capital management improving operating cash flow",
    ],

    "Order Visibility": [
        "strong order book providing revenue visibility",
        "new order inflows at record high levels",
        "robust pipeline of orders from key customers",
        "order backlog covering multiple quarters of revenue",
        "large order wins strengthening growth visibility",
    ],

    "Corporate Action": [
        "strategic acquisition adding capabilities and market access",
        "merger creating synergies and scale benefits",
        "demerger unlocking value for shareholders",
        "corporate restructuring simplifying group structure",
        "acquisition of complementary business expanding addressable market",
    ],
}


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 3 ─ NEGATIVE SENTIMENT SIGNAL LIBRARY (regex fallback)
# ─────────────────────────────────────────────────────────────────────────────

# These patterns are used ONLY when the DistilBERT model cannot be loaded.
_NEG_SIGNAL_PATTERNS: list[str] = [
    r"\b(?:delay(?:ed?|s|ing)?|postpone(?:d|s|ing)?|setback)\b",
    r"\b(?:declin(?:ed?|ing|e)?|decreas(?:ed?|ing|e)?|fell?|fall(?:ing)?|drop(?:ped?|ping)?)\b",
    r"\b(?:challeng(?:ed?|ing|es?)|difficult(?:y|ies)?|headwind[s]?|pressure[sd]?)\b",
    r"\b(?:below\s+(?:expectation|guidance|target)|miss(?:ed?|ing)?|shortfall|underperform)\b",
    r"\b(?:took\s+a\s+hit|margin\s+compression|erosion|adversely|negatively\s+impact)\b",
    r"\b(?:not\s+able|unable|couldn.t|could\s+not|failed?\s+to|did\s+not\s+achiev)\b",
    r"\b(?:concerns?|uncertainty|uncertainties|cautious|subdued|muted)\b",
    r"\b(?:slowdown|weakness|weak(?:ened?)?|soft(?:ness)?|sluggish)\b",
    r"\b(?:elevated\s+(?:cost[s]?|rm|input|raw\s+material))\b",
    r"\b(?:impact(?:ed)?\s+by|affected\s+by|hurt\s+by|weighed\s+by|dragged\s+by)\b",
]

# These patterns alone are enough to discard a sentence regardless of other signals
_STRONG_NEG_OVERRIDES: list[str] = [
    r"\b(?:delay(?:ed?|s|ing)?|postpone(?:d|s|ing)?)\b",
    r"\btook\s+a\s+hit\b",
    r"\bheadwind[s]?\b",
    r"\belevated\s+(?:cost[s]?|rm|input|raw\s+material)\b",
    r"\bbelow\s+(?:expectation[s]?|guidance)\b",
    r"\bmargin\s+compression\b",
    r"\badversely\s+impact\b",
    r"\b(?:halt(?:ed|s|ing)?|pause[sd]?|stop(?:ped)?)\b",
]

# Pre-compile for speed
_NEG_COMPILED = [re.compile(p) for p in _NEG_SIGNAL_PATTERNS]
_STRONG_NEG_COMPILED = [re.compile(p) for p in _STRONG_NEG_OVERRIDES]


# ─────────────────────────────────────────────────────────────────────────────
# MODEL INITIALISATION  ─ call nlp_engine_init() ONCE at startup
# ─────────────────────────────────────────────────────────────────────────────

# Module-level singletons so models are loaded only once per process
_spacy_nlp = None           # spaCy pipeline (sentence tokenizer)
_semantic_model = None      # SentenceTransformer
_concept_embeddings = None  # Pre-computed embeddings for SEMANTIC_CONCEPTS
_sentiment_pipeline = None  # HuggingFace sentiment classifier
_use_regex_sentiment = False  # Fallback flag


def nlp_engine_init() -> None:
    """
    Load all ML models into module-level singletons.
    Call this ONCE at the top of main() or demo_mode() before the scoring loop.

    Models downloaded on first run (~80 MB for MiniLM + ~250 MB for DistilBERT).
    Subsequent runs load from local cache — no internet required.
    """
    global _spacy_nlp, _semantic_model, _concept_embeddings
    global _sentiment_pipeline, _use_regex_sentiment

    # ── spaCy sentence tokeniser ─────────────────────────────────────────────
    log.info("NLP Engine: loading spaCy en_core_web_sm …")
    try:
        import spacy
        _spacy_nlp = spacy.load("en_core_web_sm", disable=["ner", "tagger", "lemmatizer"])
        # Increase max length for long transcripts
        _spacy_nlp.max_length = 2_000_000
        log.info("  ✓ spaCy loaded")
    except Exception as e:
        log.warning(f"  spaCy not available ({e}). Using fallback sentence splitter.")
        _spacy_nlp = None

    # ── SentenceTransformers (Layer 2 — semantic search) ─────────────────────
    log.info("NLP Engine: loading SentenceTransformer all-MiniLM-L6-v2 …")
    try:
        from sentence_transformers import SentenceTransformer
        _semantic_model = SentenceTransformer("all-MiniLM-L6-v2")
        # Pre-compute embeddings for all concept phrases (done once, reused per transcript)
        all_concepts = []
        concept_category_index = []
        for cat, phrases in SEMANTIC_CONCEPTS.items():
            for phrase in phrases:
                all_concepts.append(phrase)
                concept_category_index.append(cat)
        concept_vecs = _semantic_model.encode(all_concepts, normalize_embeddings=True,
                                              show_progress_bar=False)
        _concept_embeddings = (concept_vecs, concept_category_index)
        log.info(f"  ✓ SentenceTransformer loaded | {len(all_concepts)} concept anchors encoded")
    except Exception as e:
        log.warning(f"  SentenceTransformer not available ({e}). Semantic search disabled.")
        _semantic_model = None
        _concept_embeddings = None

    # ── HuggingFace Sentiment (Layer 3 — tone detection) ─────────────────────
    log.info("NLP Engine: loading DistilBERT sentiment classifier …")
    try:
        from transformers import pipeline as hf_pipeline
        _sentiment_pipeline = hf_pipeline(
            "sentiment-analysis",
            model="distilbert-base-uncased-finetuned-sst-2-english",
            device=-1,          # CPU only; change to 0 for GPU if available
            truncation=True,
            max_length=128,
        )
        _use_regex_sentiment = False
        log.info("  ✓ DistilBERT sentiment classifier loaded")
    except Exception as e:
        log.warning(f"  DistilBERT not available ({e}). Using regex sentiment fallback.")
        _sentiment_pipeline = None
        _use_regex_sentiment = True

    log.info("NLP Engine: initialisation complete\n")


# ─────────────────────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _split_sentences(text: str) -> list[str]:
    """
    Split transcript text into individual sentences.
    Uses spaCy if loaded, otherwise falls back to a reliable regex splitter.
    """
    if _spacy_nlp is not None:
        try:
            doc = _spacy_nlp(text)
            return [sent.text.strip() for sent in doc.sents
                    if len(sent.text.strip()) >= MIN_SENTENCE_LEN]
        except Exception:
            pass  # fall through to regex splitter

    # Regex fallback: split on ". ", "! ", "? " but keep common abbreviations intact
    text_flat = re.sub(r"[\r\n\t]+", " ", text)
    # Protect common abbreviations
    text_flat = re.sub(r"\b(Mr|Mrs|Ms|Dr|vs|etc|approx|Co|Ltd|Pvt|INR|Rs|FY|Q[1-4])\.", r"\1<DOT>", text_flat)
    # Split
    raw_sents = re.split(r"(?<=[.!?])\s+", text_flat)
    # Restore dots
    sents = [s.replace("<DOT>", ".").strip() for s in raw_sents]
    return [s for s in sents if len(s) >= MIN_SENTENCE_LEN]


def _extract_percentage_from_match(m: re.Match) -> Optional[float]:
    """Safely extract and return the captured percentage from a regex match."""
    try:
        return float(m.group(1))
    except (IndexError, ValueError, TypeError):
        return None


def _layer1_quantitative(sentence: str) -> tuple[Optional[str], Optional[float], Optional[str]]:
    """
    Layer 1: Scan a sentence for a metric-attached percentage ≥ QUANTITATIVE_MIN_PCT.

    Returns:
        (category, percentage, trigger_label) if found, else (None, None, None)
    """
    s = sentence.lower()
    # Normalise number formats: "25%" / "25 %" / "25 per cent" / "25 percent"
    s = re.sub(r"(\d+(?:\.\d+)?)\s*per\s*cent", r"\1%", s)

    for category, patterns in QUANTITATIVE_PATTERNS.items():
        for pat in patterns:
            m = re.search(pat, s)
            if m:
                pct = _extract_percentage_from_match(m)
                if pct is None:
                    continue
                
                # BPS logic remains the same
                if "bps" in pat:
                    return (category, pct, f"{int(pct)} bps")
                
                # The 20% Rule ONLY applies to Topline Growth
                if category == "High Growth / Topline":
                    if pct >= QUANTITATIVE_MIN_PCT:
                        return (category, pct, f"Growth ≥ {pct:.0f}%")
                    else:
                        continue # Silently drop it if revenue growth is under 20%
                else:
                    # For Margins, Market Share, etc., accept any extracted percentage
                    return (category, pct, f"{pct:.0f}%")
    return (None, None, None)


def _layer2_semantic(sentence: str) -> Optional[str]:
    """
    Layer 2: Compute cosine similarity between the sentence embedding and
    pre-computed concept embeddings.  Return the best-matching category
    if similarity exceeds SEMANTIC_SIMILARITY_THRESHOLD, else None.

    Returns None if SentenceTransformers is not initialised.
    """
    if _semantic_model is None or _concept_embeddings is None:
        return None

    concept_vecs, concept_category_index = _concept_embeddings

    try:
        sent_vec = _semantic_model.encode([sentence], normalize_embeddings=True,
                                          show_progress_bar=False)[0]
        # Cosine similarity = dot product of unit vectors
        sims = concept_vecs @ sent_vec
        best_idx = int(np.argmax(sims))
        best_score = float(sims[best_idx])
        if best_score >= SEMANTIC_SIMILARITY_THRESHOLD:
            return concept_category_index[best_idx]
    except Exception:
        pass

    return None


def _layer3_is_negative(sentence: str) -> bool:
    """
    Layer 3: Return True if the sentence has negative/headwind tone.

    Special case: "despite headwinds / notwithstanding challenges" is POSITIVE
    framing — the company overcame the obstacle. These are checked first and
    short-circuit the negative detection entirely.
    """
    s = sentence.lower()

    # ── 0. CONCESSION GUARD ───────────────────────────────────────────────────
    # "Revenue grew 22% despite headwinds" is positive — the company overcame it.
    # If a concession word immediately precedes or follows a negative word,
    # treat the whole sentence as positive and skip all further checks.
    _DESPITE_PAT = re.compile(
        r"\b(?:despite|notwithstanding|in\s+spite\s+of)\b[\w\s,;]{0,60}"
        r"(?:headwind|challenge|difficult|pressure|setback|weak|adverse|macro)",
        re.I,
    )
    _DESPITE_REV_PAT = re.compile(
        r"\b(?:headwind|challenge|difficult|pressure|setback|weak|adverse|macro)"
        r"[\w\s,;]{0,60}\b(?:despite|notwithstanding|in\s+spite\s+of)\b",
        re.I,
    )
    if _DESPITE_PAT.search(sentence) or _DESPITE_REV_PAT.search(sentence):
        return False  # positive framing — keep this sentence

    # ── 1. ABSOLUTE DEALBREAKERS (Runs First) ─────────────────────────────────
    # If the CEO mentions a delay, halt, or hit, we instantly throw it in the 
    # trash, even if they try to spin it as an "opportunity".
    if any(p.search(s) for p in _STRONG_NEG_COMPILED):
        return True

    # ── 2. DistilBERT AI Path (For subtler negativity) ────────────────────────
    if _sentiment_pipeline is not None and not _use_regex_sentiment:
        try:
            result = _sentiment_pipeline(sentence[:512])[0]
            if result["label"] == "NEGATIVE" and result["score"] >= SENTIMENT_NEGATIVE_THRESHOLD:
                return True
            return False
        except Exception:
            pass  # fall through to regex fallback

    # ── 3. Regex Fallback (For minor soft signals) ────────────────────────────
    soft_hits = sum(1 for p in _NEG_COMPILED if p.search(s))
    return soft_hits >= 2

def _extract_all_percentages(sentence: str) -> list[float]:
    """Return every percentage number in a sentence (handles 'per cent' spelling too)."""
    s = re.sub(r"(\d+(?:\.\d+)?)\s*per\s*cent", r"\1%", sentence.lower())
    return [float(x) for x in re.findall(r"(\d+(?:\.\d+)?)\s*%", s)]


def _also_keyword_matched(sentence: str) -> tuple[Optional[str], Optional[str]]:
    """
    Run the original KEYWORDS_DICT check on a single sentence.
    Returns (category, trigger_word) if matched, else (None, None).

    20% GATE: For 'High Growth / Topline', if the sentence explicitly states
    a percentage AND that percentage is below QUANTITATIVE_MIN_PCT, the match
    is discarded. This closes the leak where sub-20% revenue sentences bypass
    Layer 1 and get scored via the keyword fallback with no threshold check.

    Rule:
      - "revenue growth of 12%"  → keyword matches, 12 < 20 → DISCARDED
      - "revenue growth of 25%"  → keyword matches, 25 >= 20 → KEPT
      - "revenue growth momentum" → keyword matches, no % stated → KEPT
        (purely qualitative claim — no number to fail the threshold)
    """
    s_lower = sentence.lower()
    search_text = re.sub(r"[^a-z0-9 \-]", " ", s_lower)
    for category, synonyms in KEYWORDS_DICT.items():
        for synonym in synonyms:
            pat = re.compile(r"(?<![a-z])" + re.escape(synonym) + r"(?![a-z])")
            if pat.search(search_text):
                # 20% gate: only applies to the revenue growth category
                if category == "High Growth / Topline":
                    pcts = _extract_all_percentages(sentence)
                    if pcts and max(pcts) < QUANTITATIVE_MIN_PCT:
                        # A % is explicitly stated and it's below threshold — discard
                        return (None, None)
                    # No % stated (qualitative) or max % >= threshold → allow
                return (category, synonym)
    return (None, None)


# ─────────────────────────────────────────────────────────────────────────────
# PRIMARY PUBLIC FUNCTION  —  drop-in replacement for score_transcript()
# ─────────────────────────────────────────────────────────────────────────────

def score_transcript(raw_text: str) -> dict:
    """
    Three-layer NLP scoring engine.  Drop-in replacement for the original
    regex-only score_transcript() — returns the exact same dict structure.

    Returns
    -------
    dict with keys:
        total_score      (int)   — number of distinct categories matched
        categories_found (list)  — list of matched category names
        words_triggered  (list)  — list of trigger phrases / match labels
        priority         (str)   — "High" / "Medium" / "Low" / "Filtered Out"
        context_snippets (str)   — bullet-point formatted matched sentences

    How it works (per sentence)
    ---------------------------
    1. LAYER 1 — Quantitative check: does the sentence contain a recognised
       metric with a number ≥ QUANTITATIVE_MIN_PCT attached?
       → If YES, provisionally flag the sentence with the category.

    2. LAYER 2 — Semantic check (only for sentences that DIDN'T match Layer 1):
       compute cosine similarity to concept anchors.
       → If similarity ≥ SEMANTIC_SIMILARITY_THRESHOLD, provisionally flag.

    3. KEYWORD fallback: for any sentence not caught by Layer 1 or 2, apply
       the original KEYWORDS_DICT regex matching for full backward coverage.

    4. LAYER 3 — Sentiment gate: any provisionally flagged sentence is
       DISCARDED if it carries negative/headwind tone.

    5. Deduplicate categories and build the output dict.
    """
    # ── Pre-process text ──────────────────────────────────────────────────────
    # Preserve punctuation for sentence splitting (do not strip dots)
    text_clean = re.sub(r"[\r\n\t]+", " ", raw_text)
    text_clean = re.sub(r"\s{2,}", " ", text_clean).strip()

    sentences = _split_sentences(text_clean)
    log.debug(f"  Scoring {len(sentences)} sentences …")

    categories_found: set[str] = set()
    words_triggered:  set[str] = set()
    extracted_quotes: list[str] = []

    # ── Per-sentence scoring loop ─────────────────────────────────────────────
    for sentence in sentences:
        provisional_category: Optional[str] = None
        trigger_label: Optional[str] = None
        match_source: str = ""

        # ── LAYER 1: Quantitative ─────────────────────────────────────────────
        cat_l1, pct_l1, label_l1 = _layer1_quantitative(sentence)
        if cat_l1 is not None:
            provisional_category = cat_l1
            trigger_label = f"Exact Metric: {label_l1}"
            match_source = "Metric Match"

        # ── LAYER 2: Semantic (only if Layer 1 did not match) ─────────────────
        if provisional_category is None:
            cat_l2 = _layer2_semantic(sentence)
            if cat_l2 is not None:
                # 20% gate: if the semantic match landed on the revenue category
                # AND the sentence explicitly states a % below the threshold, discard.
                if cat_l2 == "High Growth / Topline":
                    pcts = _extract_all_percentages(sentence)
                    if pcts and max(pcts) < QUANTITATIVE_MIN_PCT:
                        cat_l2 = None  # stated % is too low — reject this match
                if cat_l2 is not None:
                    provisional_category = cat_l2
                    trigger_label = "Conceptual Theme"
                    match_source = "Contextual Match"

        # ── KEYWORD FALLBACK (Now completely merged into Contextual Match) ────
        if provisional_category is None:
            cat_kw, word_kw = _also_keyword_matched(sentence)
            if cat_kw is not None:
                provisional_category = cat_kw
                # We label it as a Contextual Match and show the triggered keyword cleanly
                trigger_label = word_kw
                match_source = "Contextual Match"

        # Nothing matched — move on
        if provisional_category is None:
            continue

        # ── LAYER 3: Sentiment gate ───────────────────────────────────────────
        if _layer3_is_negative(sentence):
            log.warning(f"KILLED: {sentence[:120]}")
            continue
        # ── Commit the match ──────────────────────────────────────────────────
        categories_found.add(provisional_category)
        if trigger_label:
            words_triggered.add(trigger_label)

        quote = f"• {sentence.strip()}"
        if not quote.endswith("."):
            quote += "."
        extracted_quotes.append(quote)

    # ── Scoring & priority ────────────────────────────────────────────────────
    total = len(categories_found)

    if total >= HIGH_PRIORITY_THRESHOLD:
        priority = "High"
    elif total >= MEDIUM_PRIORITY_THRESHOLD:
        priority = "Medium"
    elif total >= LOW_PRIORITY_THRESHOLD:
        priority = "Low"
    else:
        priority = "Filtered Out"

    final_quotes = (
        "\n\n".join(extracted_quotes[:MAX_SNIPPETS])
        if extracted_quotes
        else "No positive catalyst sentences found."
    )

    return {
        "total_score":      total,
        "categories_found": list(categories_found),
        "words_triggered":  list(words_triggered),
        "priority":         priority,
        "context_snippets": final_quotes,
    }



# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — CSV EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_results_to_excel(results: list[dict], discarded: list[dict], output_path: Path):
    """
    Write analysis results and discarded companies to a multi-tab Excel file.

    Results dict now carries extra keys from the multi-concall pipeline:
        weighted_score   (float)  — 0–11 weighted across 3 years
        concall_breakdown (str)   — "Q3FY25: Cat1, Cat2 | Q2FY25: ..." per-cell text
        transcript_urls  (str)    — pipe-separated list of all PDF URLs used
    """
    if not results and not discarded:
        log.warning("No results to export.")
        return

    excel_path = output_path.with_suffix('.xlsx')

    # ── Build DataFrames ──────────────────────────────────────────────────────
    df_passed = pd.DataFrame([{
        "Ticker":                      row["symbol"],
        "Company Name":                row["company_name"],
        "Market Cap (Crores)":         round(row["market_cap_crores"], 2) if row["market_cap_crores"] else "N/A",
        "Concalls Analysed (3Y)":      row.get("concalls_analysed", 1),
        "Weighted Score (0–11)":       row.get("weighted_score", row.get("total_score", 0)),
        "Cumulative Categories":       ", ".join(sorted(row["categories_found"])),
        "Priority Status":             row["priority"],
        "Concall-wise Breakdown":      row.get("concall_breakdown", ""),
        "Key Management Quotes":       row["context_snippets"],
        "Transcript Links":            row.get("transcript_urls", row.get("transcript_url", "N/A")),
    } for row in results])

    df_discarded = pd.DataFrame([{
        "Ticker":              row["symbol"],
        "Company Name":        row["company_name"],
        "Reason for Discard":  row["reason"],
    } for row in discarded])

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Styles ────────────────────────────────────────────────────────────────
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_text  = Font(color="006100", bold=True)
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_text = Font(color="9C5700", bold=True)
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_text    = Font(color="9C0006", bold=True)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_top    = Alignment(vertical="top", wrap_text=True)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

        # ── TAB 1: Qualified Companies ────────────────────────────────────────
        if not df_passed.empty:
            df_passed.to_excel(writer, sheet_name="Qualified Companies", index=False)
            ws = writer.sheets["Qualified Companies"]

            # Style header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Column index helpers
            link_col_idx     = df_passed.columns.get_loc("Transcript Links") + 1
            priority_col_idx = df_passed.columns.get_loc("Priority Status") + 1
            wscore_col_idx   = df_passed.columns.get_loc("Weighted Score (0–11)") + 1
            breakdown_col_idx= df_passed.columns.get_loc("Concall-wise Breakdown") + 1

            for row_num in range(2, ws.max_row + 1):
                # Hyperlink: if single URL, make clickable; if multiple, show first
                link_cell = ws.cell(row=row_num, column=link_col_idx)
                raw_url   = str(link_cell.value or "")
                first_url = raw_url.split(" | ")[0].strip()
                if first_url.startswith("http"):
                    link_cell.hyperlink = first_url
                    n_links = len([u for u in raw_url.split(" | ") if u.strip().startswith("http")])
                    link_cell.value = f"View PDFs ({n_links})" if n_links > 1 else "View PDF"
                    link_cell.font  = Font(color="0563C1", underline="single")

                # Priority colour
                p_cell = ws.cell(row=row_num, column=priority_col_idx)
                if p_cell.value == "High":
                    p_cell.fill = green_fill;  p_cell.font = green_text
                elif p_cell.value == "Medium":
                    p_cell.fill = yellow_fill; p_cell.font = yellow_text
                elif p_cell.value == "Low":
                    p_cell.fill = red_fill;    p_cell.font = red_text

                # Weighted score: bold the number
                ws.cell(row=row_num, column=wscore_col_idx).font = Font(bold=True)

            # Auto-width + wrap all cells
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    cell.alignment = wrap_top
                    try:
                        val_str = str(cell.value) if cell.value else ""
                        # For breakdown column, measure first pipe-segment only
                        if col_idx == breakdown_col_idx:
                            val_str = val_str.split(" | ")[0]
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 70)

            # Breakdown column: a bit wider to be readable
            ws.column_dimensions[get_column_letter(breakdown_col_idx)].width = 65
            # Freeze top row
            ws.freeze_panes = "A2"

        # ── TAB 2: Discarded Log ──────────────────────────────────────────────
        if not df_discarded.empty:
            df_discarded.to_excel(writer, sheet_name="Discarded Log", index=False)
            dws = writer.sheets["Discarded Log"]
            for cell in dws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_idx, col in enumerate(dws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    cell.alignment = wrap_top
                    try:
                        if cell.value and len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except Exception:
                        pass
                dws.column_dimensions[col_letter].width = min(max_len + 2, 55)
            dws.freeze_panes = "A2"

    log.info(f"\n✅ Dashboard generated → {excel_path}")
# ─────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # ── Setup directories ────────────────────────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TRANSCRIPTS_DIR.mkdir(parents=True, exist_ok=True)

    # NLP models loaded once before the scoring loop
    nlp_engine_init()

    # Re-attach file log handler after dirs exist
    fh = logging.FileHandler(OUTPUT_DIR / "run.log")
    fh.setFormatter(logging.Formatter("%(asctime)s  [%(levelname)s]  %(message)s", "%H:%M:%S"))
    log.addHandler(fh)

    log.info("=" * 65)
    log.info("  CONCALL TRANSCRIPT ANALYZER  — Full NSE Universe | 3-Year Window")
    log.info(f"  Weighting: Y1={YEAR_WEIGHT_Y1*100:.0f}%  Y2={YEAR_WEIGHT_Y2*100:.0f}%  Y3={YEAR_WEIGHT_Y3*100:.0f}%")
    log.info("=" * 65)

    # ── 1. Universe (Nifty 50 test mode) ─────────────────────────────────────
    # QUANTITATIVE_MIN_PCT is 20.0 by default — correct for full NSE.
    # For Nifty 50 test (large-caps grow slower) we lower it to 10.0.
    global QUANTITATIVE_MIN_PCT
    QUANTITATIVE_MIN_PCT = 20.0
    log.info(f"  Revenue growth gate: ≥{QUANTITATIVE_MIN_PCT:.0f}% (Nifty 50 test mode)")

    mcap_df = pd.read_csv("market_cap_data.csv")
    mcap_df.columns = mcap_df.columns.str.strip()
    df = pd.DataFrame({
        "symbol":            mcap_df["NSE Code"].str.strip(),
        "company_name":      mcap_df["Name"].str.strip(),
        "market_cap_crores": mcap_df["Market Capitalization"],
    })
    df = df.dropna(subset=["symbol"])
    df = df[df["symbol"].str.strip() != ""]
    log.info(f"Universe loaded from CSV: {len(df)} companies")
    if df.empty:
        log.error("Could not load universe from CSV. Exiting.")
        return

    # ── 2. Filter pipeline ───────────────────────────────────────────────────
    df_filtered = run_filter_pipeline(df)
    if df_filtered.empty:
        log.error("Zero companies survived the filter pipeline. Exiting.")
        return

    # ── 3. Phase A: Multi-concall URL resolution + download (sequential) ─────
    fetcher = ScreenerTranscriptFetcher()

    results   = []
    discarded = []

    # downloaded_jobs: one entry PER CONCALL (not per company)
    # Each entry: {symbol, company_name, market_cap_crores, concall_date, pdf_url, pdf_path}
    downloaded_jobs: list[dict] = []

    # Load completed symbols checkpoint if exists
    CHECKPOINT_FILE = OUTPUT_DIR / "completed_symbols.json"
    completed_symbols = set()
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE) as f:
            completed_symbols = set(json.load(f))
        log.info(f"  Resuming — {len(completed_symbols)} companies already completed")

    for idx, row in df_filtered.iterrows():
        symbol = row["symbol"]
        name   = row["company_name"]
        mcap   = row["market_cap_crores"]

        log.info(f"\n[{idx+1}/{len(df_filtered)}]  {symbol}  ({name})")

        if symbol in completed_symbols:
            log.info(f"  [{symbol}] Already scored — skipping")
            continue
        
        # ── Collect all transcript URLs for this company ──────────────────────
        url_list: list[tuple["datetime", str]] = []

        # Primary: Screener (returns list of (date, url))
        screener_links = fetcher.fetch_all_transcript_urls(symbol)
        url_list.extend(screener_links)

        # If Screener returned nothing, go straight to NSE
        if not url_list:
            log.info(f"  [{symbol}] Screener returned 0 links — trying NSE …")
            nse_links = fetcher.fetch_all_from_nse(symbol)
            url_list.extend(nse_links)

        if not url_list:
            log.info(f"  [{symbol}] No transcript PDFs found on Screener or NSE — skipping")
            discarded.append({
                "symbol": symbol, "company_name": name,
                "reason": "No concall PDFs found (Screener + NSE, 3-year window)"
            })
            continue

        log.info(f"  [{symbol}] Total transcript links collected: {len(url_list)}")
        url_list = url_list[:12]

        # ── Download each PDF that we don't already have cached ───────────────
        company_jobs_count = 0
        for concall_date, pdf_url in url_list:
            # Build a filename that embeds the date so concalls don't overwrite each other
            date_tag  = concall_date.strftime("%Y%m") if concall_date else "unk"
            pdf_path  = TRANSCRIPTS_DIR / f"{symbol}_{date_tag}_concall.pdf"

            if pdf_path.exists():
                # Already cached — validate before trusting
                try:
                    with open(pdf_path, "rb") as f:
                        raw = f.read(1024)
                    if raw.lstrip()[:5] == b"%PDF-" and pdf_path.stat().st_size >= 5_120:
                        log.info(f"  [{symbol}] Using cached PDF: {pdf_path.name}")
                        downloaded_jobs.append({
                            "symbol":            symbol,
                            "company_name":      name,
                            "market_cap_crores": mcap,
                            "concall_date":      concall_date,
                            "pdf_url":           pdf_url,
                            "pdf_path":          pdf_path,
                        })
                        company_jobs_count += 1
                        continue
                    else:
                        log.warning(f"  [{symbol}] Cached PDF invalid — re-downloading")
                        pdf_path.unlink(missing_ok=True)
                except Exception:
                    pdf_path.unlink(missing_ok=True)

            # Download fresh
            ok = fetcher.download_pdf(pdf_url, pdf_path)
            if not ok:
                # BSE/Screener blocked — try NSE for this specific period
                log.info(f"  [{symbol}] Download failed for {date_tag} — checking NSE …")
                nse_links = fetcher.fetch_all_from_nse(symbol)
                # Pick the NSE link closest in time to this concall_date
                best_nse = None
                best_delta = None
                for nse_dt, nse_url in nse_links:
                    delta = abs((nse_dt - concall_date).days)
                    if best_delta is None or delta < best_delta:
                        best_delta = delta
                        best_nse = (nse_dt, nse_url)
                if best_nse and best_delta is not None and best_delta <= 45:
                    nse_dt, nse_url = best_nse
                    ok2 = fetcher.download_pdf(nse_url, pdf_path)
                    if ok2:
                        pdf_url = nse_url
                        log.info(f"  [{symbol}] NSE fallback download OK ({date_tag})")
                    else:
                        log.warning(f"  [{symbol}] NSE fallback also failed for {date_tag} — skipping this concall")
                        continue
                else:
                    log.warning(f"  [{symbol}] No matching NSE link for {date_tag} — skipping this concall")
                    continue

            _polite_sleep()

            # Validate downloaded file
            try:
                with open(pdf_path, "rb") as f:
                    raw = f.read(1024)
                if raw.lstrip()[:5] != b"%PDF-" or pdf_path.stat().st_size < 5_120:
                    log.warning(f"  [{symbol}] Invalid PDF after download ({date_tag}) — skipping")
                    pdf_path.unlink(missing_ok=True)
                    continue
            except Exception:
                continue

            downloaded_jobs.append({
                "symbol":            symbol,
                "company_name":      name,
                "market_cap_crores": mcap,
                "concall_date":      concall_date,
                "pdf_url":           pdf_url,
                "pdf_path":          pdf_path,
            })
            company_jobs_count += 1

        if company_jobs_count == 0:
            log.warning(f"  [{symbol}] All downloads failed — discarding company")
            discarded.append({
                "symbol": symbol, "company_name": name,
                "reason": "All PDF downloads failed across 3-year window"
            })

        if (idx + 1) % 50 == 0:
            log.info(f"  Checkpoint: {idx+1} companies processed so far")
            export_results_to_excel(results, discarded, OUTPUT_DIR / "concall_analysis_checkpoint")

        completed_symbols.add(symbol)
        with open(CHECKPOINT_FILE, "w") as f:
            json.dump(list(completed_symbols), f)
            
        if (idx + 1) % 50 == 0:
            log.info(f"  Checkpoint: {idx+1} companies processed so far")
            export_results_to_excel(results, discarded, OUTPUT_DIR / "concall_analysis_checkpoint")

    log.info(f"\n  Download phase complete. {len(downloaded_jobs)} PDFs queued for parsing.\n")

    # ── 4. Phase B: Parse + score in parallel ────────────────────────────────
    PDF_PARSE_WORKERS = 8

    def _parse_and_score_concall(job: dict) -> dict:
        """Score a single concall PDF. Returns result dict or discard dict."""
        sym       = job["symbol"]
        name_     = job["company_name"]
        mcap_     = job["market_cap_crores"]
        pdf_url_  = job["pdf_url"]
        pdf_path_ = job["pdf_path"]
        cc_date   = job["concall_date"]

        text = extract_text_from_pdf(pdf_path_)
        if not text.strip():
            log.warning(f"  [{sym}] Empty text from {pdf_path_.name}")
            return {"_discard": True, "symbol": sym, "company_name": name_,
                    "reason": f"PDF text extraction failed ({pdf_path_.name})"}

        score = score_transcript(text)
        log.info(f"  [{sym}] {_fy_label(cc_date)} → score={score['total_score']} "
                 f"priority={score['priority']} "
                 f"cats={', '.join(score['categories_found']) or 'none'}")

        return {
            "_discard":          False,
            "symbol":            sym,
            "company_name":      name_,
            "market_cap_crores": mcap_,
            "concall_date":      cc_date,
            "pdf_url":           pdf_url_,
            **score,
        }

    import concurrent.futures
    per_concall_results: list[dict] = []
    per_concall_discards: list[dict] = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=PDF_PARSE_WORKERS) as executor:
        futures = [executor.submit(_parse_and_score_concall, job) for job in downloaded_jobs]
        for future in concurrent.futures.as_completed(futures):
            try:
                outcome = future.result()
                if outcome.get("_discard"):
                    per_concall_discards.append({
                        "symbol":       outcome["symbol"],
                        "company_name": outcome["company_name"],
                        "reason":       outcome["reason"],
                    })
                else:
                    outcome.pop("_discard", None)
                    per_concall_results.append(outcome)
            except Exception as e:
                log.warning(f"  Parse/score worker error: {e}")

    # ── 5. Aggregate per-company ──────────────────────────────────────────────
    # Group individual concall scores → one row per company
    from collections import defaultdict

    company_buckets: dict[str, list[dict]] = defaultdict(list)
    for rec in per_concall_results:
        company_buckets[rec["symbol"]].append(rec)

    # Companies where every concall failed scoring get discarded
    symbols_with_results = set(company_buckets.keys())
    symbols_attempted    = {j["symbol"] for j in downloaded_jobs}
    for sym in symbols_attempted - symbols_with_results:
        # Check if already discarded at download stage
        already = any(d["symbol"] == sym for d in discarded)
        if not already:
            # Get company name from downloaded_jobs
            name_ = next((j["company_name"] for j in downloaded_jobs if j["symbol"] == sym), sym)
            discarded.append({
                "symbol": sym, "company_name": name_,
                "reason": "All concalls scored 0 categories after NLP filtering"
            })

    for sym, concall_list in company_buckets.items():
        # Sort newest-first for breakdown display
        concall_list.sort(key=lambda r: r["concall_date"], reverse=True)

        # ── Cumulative categories (union across all concalls) ─────────────────
        all_categories: set[str] = set()
        for rec in concall_list:
            all_categories.update(rec["categories_found"])

        # ── Weighted score ────────────────────────────────────────────────────
        weighted = compute_weighted_score([
            {"date": r["concall_date"], "total_score": r["total_score"]}
            for r in concall_list
        ])

        # ── Priority from weighted score ──────────────────────────────────────
        if weighted >= HIGH_PRIORITY_THRESHOLD:
            priority = "High"
        elif weighted >= MEDIUM_PRIORITY_THRESHOLD:
            priority = "Medium"
        elif weighted >= LOW_PRIORITY_THRESHOLD:
            priority = "Low"
        else:
            priority = "Filtered Out"

        # ── Concall-wise breakdown string ─────────────────────────────────────
        # Format: "Q3 FY25: Order Visibility, Margin Expansion | Q2 FY25: ..."
        breakdown_parts = []
        for rec in concall_list:
            label = _fy_label(rec["concall_date"])
            cats  = ", ".join(sorted(rec["categories_found"])) if rec["categories_found"] else "No catalysts"
            breakdown_parts.append(f"{label}: {cats}")
        breakdown_str = "\n".join(breakdown_parts)

        # ── Transcript links (pipe-separated) ────────────────────────────────
        urls_str = " | ".join(r["pdf_url"] for r in concall_list)

        # ── Best context snippets (from most recent concall with snippets) ────
        best_snippets_parts = []
        for rec in concall_list:
            if rec.get("context_snippets") and "No positive" not in rec["context_snippets"]:
                label = _fy_label(rec["concall_date"])
                best_snippets_parts.append(f"[{label}]\n{rec['context_snippets']}")
        best_snippets = "\n\n".join(best_snippets_parts) if best_snippets_parts else "No positive catalyst sentences found."

        if priority == "Filtered Out":
            log.info(f"  [{sym}] Weighted score {weighted:.2f} → below threshold, discarding")
            name_ = concall_list[0]["company_name"]
            discarded.append({
                "symbol": sym, "company_name": name_,
                "reason": f"Weighted score {weighted:.2f} below Low threshold ({LOW_PRIORITY_THRESHOLD})"
            })
            continue

        results.append({
            "symbol":            sym,
            "company_name":      concall_list[0]["company_name"],
            "market_cap_crores": concall_list[0]["market_cap_crores"],
            "concalls_analysed": len(concall_list),
            "weighted_score":    weighted,
            "total_score":       len(all_categories),        # cumulative distinct categories
            "categories_found":  list(all_categories),
            "words_triggered":   [],                         # not aggregated at company level
            "priority":          priority,
            "concall_breakdown": breakdown_str,
            "context_snippets":  best_snippets,
            "transcript_urls":   urls_str,
        })

    # ── 6. Export ─────────────────────────────────────────────────────────────
    # Merge parse-level discards into main discard list (dedup by symbol)
    seen_discard_syms = {d["symbol"] for d in discarded}
    for d in per_concall_discards:
        if d["symbol"] not in seen_discard_syms:
            discarded.append(d)
            seen_discard_syms.add(d["symbol"])

    OUTPUT_EXCEL = OUTPUT_DIR / f"concall_analysis_3Y_{datetime.today().strftime('%Y%m%d_%H%M%S')}"
    export_results_to_excel(results, discarded, OUTPUT_EXCEL)

    # ── 7. Run summary ────────────────────────────────────────────────────────
    log.info(f"\n{'='*65}")
    log.info(f"  Summary  (Nifty 50 | 3-Year Window | Revenue gate ≥{QUANTITATIVE_MIN_PCT:.0f}%)")
    log.info(f"{'='*65}")
    log.info(f"  Companies in universe        : {len(df)}")
    log.info(f"  After market cap filter      : {len(df_filtered)}")
    log.info(f"  Total PDFs downloaded        : {len(downloaded_jobs)}")
    log.info(f"  Total concalls scored        : {len(per_concall_results)}")
    log.info(f"  Companies in final output    : {len(results)}")
    high = sum(1 for r in results if r["priority"] == "High")
    med  = sum(1 for r in results if r["priority"] == "Medium")
    low  = sum(1 for r in results if r["priority"] == "Low")
    log.info(f"    High priority              : {high}")
    log.info(f"    Medium priority            : {med}")
    log.info(f"    Low priority               : {low}")
    log.info(f"  Discarded                    : {len(discarded)}")
    avg_concalls = (sum(r["concalls_analysed"] for r in results) / len(results)) if results else 0
    log.info(f"  Avg concalls/company (output): {avg_concalls:.1f}")
    log.info(f"  Output Excel                 : {OUTPUT_EXCEL}.xlsx")
    log.info("=" * 65)


# ─────────────────────────────────────────────────────────────────────────────
# OFFLINE / DEMO MODE
# ─────────────────────────────────────────────────────────────────────────────
# When you want to test the pipeline without live HTTP calls, run:
#   python concall_analyzer.py --demo
#
# This feeds synthetic data through the scoring engine and exports a sample CSV.
# ─────────────────────────────────────────────────────────────────────────────

def demo_mode():
    """
    Test the pipeline end-to-end with synthetic transcript snippets.
    No internet access required.  Simulates 3 concalls per company.
    """
    from datetime import datetime, timedelta

    log.info("=" * 65)
    log.info("  DEMO MODE — synthetic data, no HTTP calls (3-year sim)")
    log.info("=" * 65)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    nlp_engine_init()
    import logging
    log.setLevel(logging.DEBUG)

    # Simulate 3 concalls per company across 3 years
    today = datetime.today()
    synthetic = [
        {
            "symbol": "RELIANCE",
            "company_name": "Reliance Industries Ltd",
            "market_cap_crores": 1_800_000,
            "concalls": [
                {
                    "date": today - timedelta(days=60),   # Q3 FY26 — Y1
                    "transcript": (
                        "our revenue grew 28% year on year backed by strong retail and telecom performance. "
                        "ebitda margin expansion of 180 bps to 19.2% greenfield expansion in new energy. "
                        "continues on track order book remains strong with new order inflows at record levels. "
                        "free cash flow positive this quarter net cash positive on the balance sheet. "
                        "debt repayment of 12000 crores completed deleveraging continues. "
                        "working capital days reduced by 6 days cash conversion cycle improvement. "
                        "return on capital employed improved to 14.5% gaining market share in key retail categories. "
                        "capacity addition of 20 mtpa brownfield underway backward integration progressing well."
                    ),
                },
                {
                    "date": today - timedelta(days=430),  # Y2
                    "transcript": (
                        "revenue growth of 22% driven by jio platforms and retail verticals. "
                        "ebitda margins expanded 150 bps to 18.5% we are gaining market share in key categories. "
                        "capex for capacity expansion remains at 75000 crores greenfield plant on schedule. "
                        "deleveraging continues with net debt declining free cash flow positive. "
                        "order book visibility strong new orders signed this quarter. "
                        "working capital improvement of 5 days return on equity improved to 13.2%."
                    ),
                },
                {
                    "date": today - timedelta(days=800),  # Y3
                    "transcript": (
                        "topline growth of 21% in fy23 driven by retail and energy segments. "
                        "backward integration into chemicals progressing well capacity addition on track. "
                        "return on capital employed improved to 14.5% debt repayment completed ahead of schedule. "
                        "ebitda margin expansion of 120 bps order inflows remain robust."
                    ),
                },
            ],
        },
        {
            "symbol": "TATASTEEL",
            "company_name": "Tata Steel Ltd",
            "market_cap_crores": 220_000,
            "concalls": [
                {
                    "date": today - timedelta(days=90),
                    "transcript": (
                        "debottlenecking at our facilities has improved throughput capacity addition of 5 mtpa brownfield underway. "
                        "debt repayment schedules on track deleveraging continues net debt to ebitda improving. "
                        "ebitda margin expanded 200 bps to 17.3% working capital days reduced by 4 days. "
                        "order book strong with new order inflows from infrastructure sector. "
                        "return on capital employed improved to 12.8% free cash flow positive this quarter."
                    ),
                },
                {
                    "date": today - timedelta(days=460),
                    "transcript": (
                        "working capital days reduced by 8 days collection efficiency improved cash conversion cycle improvement. "
                        "net cash positive on standalone basis strong free cash flow this quarter. "
                        "capacity addition of 12000 crores capex on schedule brownfield expansion progressing. "
                        "ebitda margins expanded 100 bps gaining market share in auto and construction segments. "
                        "deleveraging on track debt repayment of 5000 crores done this year."
                    ),
                },
            ],
        },
    ]

    results   = []
    discarded = []

    for co in synthetic:
        if co["market_cap_crores"] < MIN_MARKET_CAP_CRORES:
            log.info(f"  {co['symbol']} filtered out by market cap")
            continue

        per_concall: list[dict] = []
        

        for cc in co["concalls"]:
            score = score_transcript(cc["transcript"])
            log.info(f"  {co['symbol']} {_fy_label(cc['date'])} → "
                     f"score={score['total_score']} cats={score['categories_found']}")
            per_concall.append({"date": cc["date"], **score})

        best_snippets_parts = []
        for r in per_concall:
            if r.get("context_snippets") and "No positive" not in r["context_snippets"]:
                label = _fy_label(r["date"])
                best_snippets_parts.append(f"[{label}]\n{r['context_snippets']}")
        best_snippets = "\n\n".join(best_snippets_parts) if best_snippets_parts else "No positive catalyst sentences found."

        if not per_concall:
            continue

        # Aggregate
        all_cats = set()
        for r in per_concall:
            all_cats.update(r["categories_found"])

        weighted = compute_weighted_score([
            {"date": r["date"], "total_score": r["total_score"]} for r in per_concall
        ])

        if   weighted >= HIGH_PRIORITY_THRESHOLD:   priority = "High"
        elif weighted >= MEDIUM_PRIORITY_THRESHOLD: priority = "Medium"
        elif weighted >= LOW_PRIORITY_THRESHOLD:    priority = "Low"
        else:                                       priority = "Filtered Out"

        per_concall.sort(key=lambda r: r["date"], reverse=True)
        breakdown = "\n".join(
            f"{_fy_label(r['date'])}: {', '.join(sorted(r['categories_found'])) or 'No catalysts'}"
            for r in per_concall
        )

        if priority != "Filtered Out":
            results.append({
                "symbol":            co["symbol"],
                "company_name":      co["company_name"],
                "market_cap_crores": co["market_cap_crores"],
                "concalls_analysed": len(per_concall),
                "weighted_score":    weighted,
                "total_score":       len(all_cats),
                "categories_found":  list(all_cats),
                "words_triggered":   [],
                "priority":          priority,
                "concall_breakdown": breakdown,
                "context_snippets":  best_snippets,
                "transcript_urls":   "N/A (Demo)",
            })

    export_results_to_excel(results, discarded, OUTPUT_DIR / "demo_output_3Y")


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if "--demo" in sys.argv:
        demo_mode()
    else:
        main()